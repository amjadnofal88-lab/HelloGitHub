#!/usr/bin/env python3
"""
report.py — Export statement entries to a 7-sheet RTL Arabic Excel workbook.

Usage:
    python report.py --db statements.db --out report.xlsx [--account ACCOUNT]

Sheets:
    1. ملخص          Summary totals
    2. كل القيود      All entries
    3. الشيكات        Cheque entries
    4. المرتجعات      Returned / bounced cheques
    5. التكرارات      Duplicate journal numbers
    6. المقاصة        Offset (net-zero) pairs
    7. الكراجات       Garage / workshop payments
"""

import argparse
import sqlite3
import sys
from datetime import datetime
from typing import Any

try:
    import openpyxl
    from openpyxl.styles import (
        Alignment, Font, PatternFill, Border, Side
    )
    from openpyxl.utils import get_column_letter
except ImportError:
    print("openpyxl غير مثبّت. شغّل: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

# ---------------------------------------------------------------------------
# Colours
# ---------------------------------------------------------------------------
HEADER_FILL  = PatternFill("solid", fgColor="1F497D")
ALT_FILL     = PatternFill("solid", fgColor="DCE6F1")
HEADER_FONT  = Font(bold=True, color="FFFFFF", name="Arial", size=11)
DATA_FONT    = Font(name="Arial", size=10)
THIN         = Side(style="thin", color="BFBFBF")
THIN_BORDER  = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
RTL_ALIGN    = Alignment(horizontal="right", vertical="center", readingOrder=2)
RTL_WRAP     = Alignment(horizontal="right", vertical="center", readingOrder=2, wrap_text=True)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center")

# ---------------------------------------------------------------------------
# DB helpers
# ---------------------------------------------------------------------------

def open_db(db_path: str) -> sqlite3.Connection:
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    return conn


def fetch(conn: sqlite3.Connection, sql: str, params: tuple = ()) -> list[sqlite3.Row]:
    return conn.execute(sql, params).fetchall()


# ---------------------------------------------------------------------------
# Workbook helpers
# ---------------------------------------------------------------------------

def make_wb() -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)          # remove default sheet
    return wb


def add_sheet(wb: openpyxl.Workbook, title: str) -> openpyxl.worksheet.worksheet.Worksheet:
    ws = wb.create_sheet(title=title)
    ws.sheet_view.rightToLeft = True
    return ws


def write_header(ws, columns: list[str], row: int = 1) -> None:
    for col, name in enumerate(columns, 1):
        cell = ws.cell(row=row, column=col, value=name)
        cell.fill   = HEADER_FILL
        cell.font   = HEADER_FONT
        cell.border = THIN_BORDER
        cell.alignment = RTL_ALIGN


def write_data_row(ws, row_idx: int, values: list[Any], alt: bool = False) -> None:
    fill = ALT_FILL if alt else None
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row_idx, column=col, value=val)
        cell.font   = DATA_FONT
        cell.border = THIN_BORDER
        cell.alignment = RTL_ALIGN
        if fill:
            cell.fill = fill


def auto_width(ws) -> None:
    for column_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 4, 10), 50)


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------

ENTRY_COLS = [
    "id", "source", "account", "date", "iso_date",
    "journal_no", "journal_type", "kind", "cheque_no", "note",
    "amount", "currency", "debit", "credit", "balance",
]

ENTRY_HEADERS = [
    "م", "المصدر", "الحساب", "التاريخ", "تاريخ ISO",
    "رقم القيد", "نوع القيد", "النوع", "رقم الشيك", "البيان",
    "المبلغ", "العملة", "المدين", "الدائن", "الرصيد",
]


def _entry_values(row: sqlite3.Row) -> list[Any]:
    return [row[c] for c in ENTRY_COLS]


def build_summary(wb, conn, where: str, params: tuple) -> None:
    ws = add_sheet(wb, "ملخص")
    cols = ["الحساب", "العملة", "عدد الصفوف", "مجموع المدين", "مجموع الدائن", "من تاريخ", "إلى تاريخ"]
    write_header(ws, cols)
    rows = fetch(conn, f"""
        SELECT account, currency,
               COUNT(*) AS cnt,
               ROUND(SUM(COALESCE(debit,0)),2)  AS tot_debit,
               ROUND(SUM(COALESCE(credit,0)),2) AS tot_credit,
               MIN(iso_date), MAX(iso_date)
        FROM entries {where}
        GROUP BY account, currency
        ORDER BY account, currency
    """, params)
    for i, row in enumerate(rows, 2):
        write_data_row(ws, i, list(row), alt=(i % 2 == 0))
    auto_width(ws)


def build_all_entries(wb, conn, where: str, params: tuple) -> None:
    ws = add_sheet(wb, "كل القيود")
    write_header(ws, ENTRY_HEADERS)
    rows = fetch(conn, f"SELECT * FROM entries {where} ORDER BY iso_date", params)
    for i, row in enumerate(rows, 2):
        write_data_row(ws, i, _entry_values(row), alt=(i % 2 == 0))
    auto_width(ws)


def build_cheques(wb, conn, where: str, params: tuple) -> None:
    ws = add_sheet(wb, "الشيكات")
    write_header(ws, ENTRY_HEADERS)
    cheque_where = where.replace("WHERE 1=1", "WHERE cheque_no IS NOT NULL AND TRIM(cheque_no)!=''")
    if "WHERE 1=1" not in where:
        cheque_where = where + " AND cheque_no IS NOT NULL AND TRIM(cheque_no)!=''"
    rows = fetch(conn, f"SELECT * FROM entries {cheque_where} ORDER BY iso_date", params)
    for i, row in enumerate(rows, 2):
        write_data_row(ws, i, _entry_values(row), alt=(i % 2 == 0))
    auto_width(ws)


def build_returned(wb, conn, where: str, params: tuple) -> None:
    ws = add_sheet(wb, "المرتجعات")
    write_header(ws, ENTRY_HEADERS)
    keywords = ["مرتجع", "راجع", "returned", "bounced", "dishonoured", "dishonored"]
    like_expr = " OR ".join("LOWER(note) LIKE ?" for _ in keywords)
    kw_params = [f"%{k}%" for k in keywords]
    connector = "AND" if where.strip() != "WHERE 1=1" else "AND"
    full_where = f"{where} AND ({like_expr})"
    rows = fetch(conn, f"SELECT * FROM entries {full_where} ORDER BY iso_date", params + tuple(kw_params))
    for i, row in enumerate(rows, 2):
        write_data_row(ws, i, _entry_values(row), alt=(i % 2 == 0))
    auto_width(ws)


def build_duplicates(wb, conn, where: str, params: tuple) -> None:
    ws = add_sheet(wb, "التكرارات")
    cols = ["الحساب", "رقم القيد", "عدد التكرار", "المعرفات"]
    write_header(ws, cols)
    rows = fetch(conn, f"""
        SELECT account, journal_no, COUNT(*) AS cnt, GROUP_CONCAT(id) AS ids
        FROM entries {where} AND journal_no IS NOT NULL
        GROUP BY account, journal_no HAVING cnt > 1
        ORDER BY cnt DESC
    """, params)
    for i, row in enumerate(rows, 2):
        write_data_row(ws, i, list(row), alt=(i % 2 == 0))
    auto_width(ws)


def build_offsets(wb, conn, where: str, params: tuple) -> None:
    ws = add_sheet(wb, "المقاصة")
    cols = ["الحساب", "رقم القيد", "الصافي", "عدد الصفوف"]
    write_header(ws, cols)
    rows = fetch(conn, f"""
        SELECT account, journal_no,
               ROUND(SUM(COALESCE(debit,0)) - SUM(COALESCE(credit,0)),4) AS net,
               COUNT(*) AS cnt
        FROM entries {where} AND journal_no IS NOT NULL
        GROUP BY account, journal_no
        HAVING ABS(net) < 0.01 AND cnt >= 2
        ORDER BY account, journal_no
    """, params)
    for i, row in enumerate(rows, 2):
        write_data_row(ws, i, list(row), alt=(i % 2 == 0))
    auto_width(ws)


def build_garage(wb, conn, where: str, params: tuple) -> None:
    ws = add_sheet(wb, "الكراجات")
    write_header(ws, ENTRY_HEADERS)
    keywords = ["كراج", "ورشة", "garage", "workshop", "مركز صيانة", "بودي"]
    like_expr = " OR ".join("LOWER(note) LIKE ?" for _ in keywords)
    kw_params = [f"%{k}%" for k in keywords]
    full_where = f"{where} AND ({like_expr})"
    rows = fetch(conn, f"SELECT * FROM entries {full_where} ORDER BY iso_date", params + tuple(kw_params))
    for i, row in enumerate(rows, 2):
        write_data_row(ws, i, _entry_values(row), alt=(i % 2 == 0))
    auto_width(ws)


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description="تصدير كشوف الحسابات التأمينية إلى Excel متعدد الأوراق"
    )
    parser.add_argument("--db",      default="statements.db", help="مسار قاعدة البيانات")
    parser.add_argument("--out",     default="report.xlsx",   help="مسار ملف الإخراج")
    parser.add_argument("--account", default=None,            help="تصفية بحساب محدد")
    parser.add_argument("--from-date", dest="from_date",      help="من تاريخ (YYYY-MM-DD)")
    parser.add_argument("--to-date",   dest="to_date",        help="إلى تاريخ (YYYY-MM-DD)")
    args = parser.parse_args()

    conn = open_db(args.db)

    # Build shared WHERE clause
    conditions = ["1=1"]
    params: list[Any] = []
    if args.account:
        conditions.append("account = ?")
        params.append(args.account)
    if args.from_date:
        conditions.append("iso_date >= ?")
        params.append(args.from_date)
    if args.to_date:
        conditions.append("iso_date <= ?")
        params.append(args.to_date)

    where = "WHERE " + " AND ".join(conditions)

    wb = make_wb()
    build_summary(wb, conn, where, tuple(params))
    build_all_entries(wb, conn, where, tuple(params))
    build_cheques(wb, conn, where, tuple(params))
    build_returned(wb, conn, where, tuple(params))
    build_duplicates(wb, conn, where, tuple(params))
    build_offsets(wb, conn, where, tuple(params))
    build_garage(wb, conn, where, tuple(params))

    conn.close()

    # Add metadata sheet
    ws_meta = add_sheet(wb, "معلومات")
    ws_meta["A1"] = "تاريخ التقرير"
    ws_meta["B1"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws_meta["A2"] = "قاعدة البيانات"
    ws_meta["B2"] = args.db

    wb.save(args.out)
    print(f"[report] تم حفظ التقرير في: {args.out}")


if __name__ == "__main__":
    main()
