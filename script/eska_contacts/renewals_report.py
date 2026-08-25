#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
renewals_report.py — توليد تقرير التجديدات من ملف الذمم وبيانات التواصل.

الاستخدام:
    python renewals_report.py "ذمم.xls" --contacts contacts.xlsx
    python renewals_report.py "ذمم.xls" --contacts contacts.xlsx --out تقرير_التجديدات.xlsx

المخرجات:
    تقرير إكسل مُنسَّق يحتوي على:
    - ورقة "التجديدات"   : الوثائق المنتهية خلال نافذة التجديد مع بيانات التواصل
    - ورقة "الكل"        : جميع الوثائق مع بيانات التواصل
    - ورقة "ملخص"        : إحصائيات إجمالية
"""

import argparse
import logging
import os
import sys
from datetime import date

import pandas as pd
import openpyxl
from openpyxl.styles import (
    Alignment,
    Font,
    PatternFill,
    Border,
    Side,
)
from openpyxl.utils import get_column_letter

# ── أعمدة ملف الذمم (يجب أن تتطابق مع eska_contacts.py) ─────────────────────
COL_POLICY_NO    = 'رقم الوثيقة'
COL_CLIENT_NAME  = 'اسم العميل'
COL_EXPIRY_DATE  = 'تاريخ الانتهاء'
COL_INSURER      = 'شركة التأمين'
COL_POLICY_TYPE  = 'نوع التأمين'
COL_PREMIUM      = 'إجمالي القسط'

# أعمدة التواصل المُضافة بواسطة eska_contacts.py
COL_PHONE  = 'هاتف'
COL_MOBILE = 'جوال'
COL_EMAIL  = 'بريد إلكتروني'

# نافذة التجديد بالأيام (يجب أن تتطابق مع eska_contacts.py)
RENEWAL_DAYS = 90

# ── ألوان التنسيق ─────────────────────────────────────────────────────────────
COLOR_HEADER_BG  = '1F4E79'   # أزرق داكن
COLOR_HEADER_FG  = 'FFFFFF'   # أبيض
COLOR_ALT_ROW    = 'D6E4F0'   # أزرق فاتح (صفوف بديلة)
COLOR_URGENT_BG  = 'FFE0E0'   # أحمر فاتح (تنتهي خلال 30 يومًا)
COLOR_SOON_BG    = 'FFF3CD'   # أصفر فاتح (تنتهي خلال 31-60 يومًا)

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s  %(levelname)-8s  %(message)s',
    datefmt='%H:%M:%S',
)
logger = logging.getLogger('renewals_report')


# ── تحميل البيانات ────────────────────────────────────────────────────────────

def load_file(path: str) -> pd.DataFrame:
    """تحميل XLS أو XLSX."""
    ext = os.path.splitext(path)[1].lower()
    engine = 'xlrd' if ext == '.xls' else 'openpyxl'
    df = pd.read_excel(path, engine=engine)
    logger.info('تم تحميل %d صف من "%s"', len(df), path)
    return df


def merge_data(xls_df: pd.DataFrame, contacts_df: pd.DataFrame) -> pd.DataFrame:
    """دمج بيانات الذمم مع بيانات التواصل على أساس رقم الوثيقة."""
    contact_cols = [c for c in [COL_PHONE, COL_MOBILE, COL_EMAIL, 'ملاحظات']
                    if c in contacts_df.columns]

    if COL_POLICY_NO not in xls_df.columns or COL_POLICY_NO not in contacts_df.columns:
        logger.warning(
            'عمود "%s" مفقود في أحد الملفين — سيتم إضافة أعمدة التواصل فارغة.',
            COL_POLICY_NO,
        )
        for col in contact_cols:
            xls_df[col] = ''
        return xls_df

    merge_contacts = contacts_df[[COL_POLICY_NO] + contact_cols].copy()
    merged = xls_df.merge(merge_contacts, on=COL_POLICY_NO, how='left', suffixes=('', '_c'))
    logger.info('بعد الدمج: %d صف', len(merged))
    return merged


def split_renewals(df: pd.DataFrame):
    """
    تقسيم DataFrame إلى وثائق تجديد وبقية الوثائق.
    تُعيد (renewals_df, all_df) حيث renewals_df مرتّبة حسب أقرب تاريخ انتهاء.
    """
    if COL_EXPIRY_DATE not in df.columns:
        logger.warning('العمود "%s" مفقود — لن يتم التصفية.', COL_EXPIRY_DATE)
        return df.copy(), df.copy()

    df = df.copy()
    df[COL_EXPIRY_DATE] = pd.to_datetime(df[COL_EXPIRY_DATE], errors='coerce')
    today = pd.Timestamp(date.today())
    cutoff = today + pd.Timedelta(days=RENEWAL_DAYS)

    mask = (df[COL_EXPIRY_DATE] >= today) & (df[COL_EXPIRY_DATE] <= cutoff)
    renewals = df[mask].sort_values(COL_EXPIRY_DATE).copy()
    logger.info('وثائق التجديد: %d', len(renewals))
    return renewals, df


# ── إنشاء التقرير بإكسل ───────────────────────────────────────────────────────

def _header_style() -> tuple:
    """إعادة Font و PatternFill للرأس."""
    font = Font(bold=True, color=COLOR_HEADER_FG, name='Calibri', size=11)
    fill = PatternFill(fill_type='solid',
                       start_color=COLOR_HEADER_BG,
                       end_color=COLOR_HEADER_BG)
    align = Alignment(horizontal='center', vertical='center',
                      wrap_text=True, reading_order=2)
    return font, fill, align


def _thin_border() -> Border:
    side = Side(style='thin', color='BFBFBF')
    return Border(left=side, right=side, top=side, bottom=side)


def _days_until(expiry) -> int:
    """عدد الأيام المتبقية حتى تاريخ الانتهاء. -1 إذا لم يكن صالحًا."""
    try:
        delta = pd.Timestamp(expiry) - pd.Timestamp(date.today())
        return int(delta.days)
    except Exception:  # noqa: BLE001
        return -1


def _row_fill(days: int):
    """إعادة PatternFill بناءً على عدد الأيام المتبقية."""
    if 0 <= days <= 30:
        color = COLOR_URGENT_BG
    elif 31 <= days <= 60:
        color = COLOR_SOON_BG
    else:
        return None
    return PatternFill(fill_type='solid', start_color=color, end_color=color)


def write_sheet(ws, df: pd.DataFrame, title: str, highlight: bool = False) -> None:
    """
    كتابة DataFrame في ورقة العمل ws مع تنسيق الرأس والصفوف.

    :param highlight: إذا True، يُلوَّن الصف بناءً على قرب تاريخ الانتهاء.
    """
    ws.title = title
    ws.sheet_view.rightToLeft = True   # دعم اتجاه RTL

    header_font, header_fill, header_align = _header_style()
    border = _thin_border()

    # ── كتابة رأس الجدول ──────────────────────────────────────────────────────
    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border

    ws.row_dimensions[1].height = 28

    # ── كتابة البيانات ────────────────────────────────────────────────────────
    expiry_col_idx = None
    if COL_EXPIRY_DATE in df.columns:
        expiry_col_idx = list(df.columns).index(COL_EXPIRY_DATE)

    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        alt_fill = PatternFill(fill_type='solid',
                               start_color=COLOR_ALT_ROW,
                               end_color=COLOR_ALT_ROW) if row_idx % 2 == 0 else None

        # تحديد تلوين الصف بناءً على تاريخ الانتهاء
        row_highlight_fill = None
        if highlight and expiry_col_idx is not None:
            expiry_val = row[expiry_col_idx]
            days = _days_until(expiry_val)
            row_highlight_fill = _row_fill(days)

        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            cell.alignment = Alignment(
                horizontal='right', vertical='center', reading_order=2
            )
            if row_highlight_fill:
                cell.fill = row_highlight_fill
            elif alt_fill:
                cell.fill = alt_fill

    # ── ضبط عرض الأعمدة ───────────────────────────────────────────────────────
    for col_idx, col_name in enumerate(df.columns, start=1):
        col_letter = get_column_letter(col_idx)
        max_len = max(
            len(str(col_name)),
            df.iloc[:, col_idx - 1].astype(str).str.len().max() if len(df) > 0 else 0,
        )
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 40)

    # ── تجميد الصف الأول ──────────────────────────────────────────────────────
    ws.freeze_panes = 'A2'


def write_summary_sheet(ws, renewals_df: pd.DataFrame, all_df: pd.DataFrame) -> None:
    """كتابة ورقة الملخص الإحصائي."""
    ws.title = 'ملخص'
    ws.sheet_view.rightToLeft = True

    header_font, header_fill, header_align = _header_style()
    border = _thin_border()
    today = date.today()

    rows = [
        ('البيان', 'القيمة'),
        ('تاريخ التقرير', str(today)),
        ('إجمالي الوثائق', len(all_df)),
        ('وثائق التجديد (خلال {} يومًا)'.format(RENEWAL_DAYS), len(renewals_df)),
    ]

    # توزيع التجديدات حسب نافذة الإلحاح
    if COL_EXPIRY_DATE in renewals_df.columns:
        exp = pd.to_datetime(renewals_df[COL_EXPIRY_DATE], errors='coerce')
        today_ts = pd.Timestamp(today)
        urgent = ((exp - today_ts).dt.days.between(0, 30)).sum()
        soon   = ((exp - today_ts).dt.days.between(31, 60)).sum()
        later  = ((exp - today_ts).dt.days.between(61, RENEWAL_DAYS)).sum()
        rows += [
            ('  ↳ عاجل (0-30 يومًا) 🔴', int(urgent)),
            ('  ↳ قريب (31-60 يومًا) 🟡', int(soon)),
            ('  ↳ لاحقاً (61-{} يومًا) 🟢'.format(RENEWAL_DAYS), int(later)),
        ]

    # إحصائيات التواصل
    if COL_MOBILE in renewals_df.columns:
        has_mobile = renewals_df[COL_MOBILE].astype(str).str.strip().str.len().gt(0).sum()
        rows.append(('تجديدات بجوال متاح', int(has_mobile)))
    if COL_EMAIL in renewals_df.columns:
        has_email = renewals_df[COL_EMAIL].astype(str).str.strip().str.len().gt(0).sum()
        rows.append(('تجديدات ببريد إلكتروني متاح', int(has_email)))

    for row_idx, (label, value) in enumerate(rows, start=1):
        c1 = ws.cell(row=row_idx, column=1, value=label)
        c2 = ws.cell(row=row_idx, column=2, value=value)
        for cell in (c1, c2):
            cell.border = border
            cell.alignment = Alignment(
                horizontal='right', vertical='center', reading_order=2
            )
        if row_idx == 1:
            for cell in (c1, c2):
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align

    ws.column_dimensions['A'].width = 38
    ws.column_dimensions['B'].width = 20


def build_report(
    xls_path: str,
    contacts_path: str,
    out_path: str,
) -> None:
    """المنطق الرئيسي لبناء التقرير."""
    xls_df = load_file(xls_path)
    contacts_df = load_file(contacts_path)

    merged = merge_data(xls_df, contacts_df)
    renewals_df, all_df = split_renewals(merged)

    wb = openpyxl.Workbook()

    # ورقة التجديدات (الورقة الأولى الافتراضية)
    ws_renewals = wb.active
    write_sheet(ws_renewals, renewals_df, 'التجديدات', highlight=True)

    # ورقة الكل
    ws_all = wb.create_sheet('الكل')
    write_sheet(ws_all, all_df, 'الكل', highlight=False)

    # ورقة الملخص
    ws_summary = wb.create_sheet('ملخص')
    write_summary_sheet(ws_summary, renewals_df, all_df)

    wb.save(out_path)
    logger.info('تم حفظ التقرير في: %s', out_path)
    print('اكتمل — تم حفظ التقرير في: {}'.format(out_path))
    print('  ↳ التجديدات : {} وثيقة'.format(len(renewals_df)))
    print('  ↳ الإجمالي  : {} وثيقة'.format(len(all_df)))


# ── نقطة الدخول الرئيسية ──────────────────────────────────────────────────────

def main() -> None:
    parser = argparse.ArgumentParser(
        description='توليد تقرير التجديدات من ملف الذمم وبيانات التواصل.',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    parser.add_argument('xls', help='مسار ملف الذمم (XLS أو XLSX)')
    parser.add_argument(
        '--contacts',
        required=True,
        help='مسار ملف بيانات التواصل (الناتج من eska_contacts.py)',
    )
    parser.add_argument(
        '--out',
        default='تقرير_التجديدات.xlsx',
        help='مسار ملف التقرير (الافتراضي: تقرير_التجديدات.xlsx)',
    )
    args = parser.parse_args()

    for path, label in [(args.xls, 'ملف الذمم'), (args.contacts, 'ملف التواصل')]:
        if not os.path.isfile(path):
            sys.exit('خطأ: {} "{}" غير موجود.'.format(label, path))

    build_report(args.xls, args.contacts, args.out)


if __name__ == '__main__':
    main()
