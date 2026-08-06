#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
agents_report.py — تقرير شامل لحسابات الوكلاء من ملف الذمم.

الاستخدام:
    python agents_report.py "ذمم.xls"
    python agents_report.py "ذمم.xls" --commission 0.10 --out وكلاء.xlsx

المخرجات (أوراق Excel):
    ملخص_عام        — إجماليات جميع الوكلاء
    عمولات          — عمولة كل وثيقة لكل وكيل
    أداء_الوكلاء    — عدد وثائق / تجديدات / إجمالي أقساط لكل وكيل
    مستحقات         — المبالغ المستحقة التي لم تُحصَّل بعد لكل وكيل
    ذمم_تفصيل       — جميع وثائق الذمم مجمَّعة مع اسم الوكيل
    <اسم_وكيل>      — ورقة مستقلة لكل وكيل بوثائقه كاملة
"""

import argparse
import logging
import os
import sys
from datetime import date

import pandas as pd
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ── أعمدة ملف الذمم ──────────────────────────────────────────────────────────
# عدّل الأسماء لتتطابق مع الملف الفعلي
COL_POLICY_NO   = 'رقم الوثيقة'
COL_CLIENT_NAME = 'اسم العميل'
COL_AGENT       = 'الوكيل'          # عمود اسم الوكيل
COL_EXPIRY_DATE = 'تاريخ الانتهاء'
COL_ISSUE_DATE  = 'تاريخ الإصدار'
COL_INSURER     = 'شركة التأمين'
COL_POLICY_TYPE = 'نوع التأمين'
COL_PREMIUM     = 'إجمالي القسط'    # القسط الإجمالي
COL_COLLECTED   = 'المحصَّل'        # ما تم تحصيله (0 إن لم يوجد)
COL_BALANCE     = 'الرصيد'          # الذمة المتبقية

# نسبة العمولة الافتراضية (يمكن تجاوزها بـ --commission)
DEFAULT_COMMISSION_RATE = 0.10   # 10%

# نافذة التجديد
RENEWAL_DAYS = 90

# ── ألوان ─────────────────────────────────────────────────────────────────────
C_HEADER_BG   = '1F4E79'
C_HEADER_FG   = 'FFFFFF'
C_ALT_ROW     = 'D6E4F0'
C_TOTAL_BG    = 'E2EFDA'   # أخضر فاتح لصفوف الإجمالي
C_OVERDUE_BG  = 'FFE0E0'   # أحمر — متأخر/منتهي
C_SOON_BG     = 'FFF3CD'   # أصفر — ينتهي قريبًا

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s  %(levelname)-8s  %(message)s',
    datefmt='%H:%M:%S',
)
logger = logging.getLogger('agents_report')


# ── تحميل البيانات ─────────────────────────────────────────────────────────────

def load_file(path: str) -> pd.DataFrame:
    ext = os.path.splitext(path)[1].lower()
    engine = 'xlrd' if ext == '.xls' else 'openpyxl'
    df = pd.read_excel(path, engine=engine)
    logger.info('تم تحميل %d صف من "%s"', len(df), path)
    return df


def prepare(df: pd.DataFrame, commission_rate: float) -> pd.DataFrame:
    """تنظيف البيانات وإضافة الأعمدة المحسوبة."""
    df = df.copy()

    # التأكد من وجود عمود الوكيل
    if COL_AGENT not in df.columns:
        logger.warning('العمود "%s" غير موجود — سيُستخدم "غير محدد".', COL_AGENT)
        df[COL_AGENT] = 'غير محدد'

    # تحويل التواريخ
    for col in (COL_EXPIRY_DATE, COL_ISSUE_DATE):
        if col in df.columns:
            df[col] = pd.to_datetime(df[col], errors='coerce')

    # القسط
    if COL_PREMIUM not in df.columns:
        df[COL_PREMIUM] = 0.0
    df[COL_PREMIUM] = pd.to_numeric(df[COL_PREMIUM], errors='coerce').fillna(0.0)

    # المحصَّل والرصيد
    if COL_COLLECTED not in df.columns:
        df[COL_COLLECTED] = 0.0
    df[COL_COLLECTED] = pd.to_numeric(df[COL_COLLECTED], errors='coerce').fillna(0.0)

    if COL_BALANCE not in df.columns:
        df[COL_BALANCE] = df[COL_PREMIUM] - df[COL_COLLECTED]
    df[COL_BALANCE] = pd.to_numeric(df[COL_BALANCE], errors='coerce').fillna(0.0)

    # عمولة كل وثيقة
    df['العمولة'] = (df[COL_PREMIUM] * commission_rate).round(2)

    # حالة الوثيقة
    today = pd.Timestamp(date.today())
    if COL_EXPIRY_DATE in df.columns:
        exp = df[COL_EXPIRY_DATE]
        df['الحالة'] = 'نشطة'
        df.loc[exp < today, 'الحالة'] = 'منتهية'
        df.loc[(exp >= today) & (exp <= today + pd.Timedelta(days=RENEWAL_DAYS)),
               'الحالة'] = 'للتجديد'

    return df


# ── حسابات الوكلاء ─────────────────────────────────────────────────────────────

def calc_commissions(df: pd.DataFrame) -> pd.DataFrame:
    """عمولة كل وثيقة."""
    cols = [c for c in [COL_POLICY_NO, COL_CLIENT_NAME, COL_AGENT,
                        COL_POLICY_TYPE, COL_INSURER, COL_PREMIUM,
                        'العمولة', COL_EXPIRY_DATE, 'الحالة'] if c in df.columns]
    return df[cols].sort_values([COL_AGENT, COL_PREMIUM], ascending=[True, False])


def calc_performance(df: pd.DataFrame) -> pd.DataFrame:
    """أداء كل وكيل: عدد الوثائق / التجديدات / الإجمالي."""
    agg = {
        COL_PREMIUM: 'sum',
        'العمولة': 'sum',
    }
    counts = df.groupby(COL_AGENT).agg(
        عدد_الوثائق=(COL_POLICY_NO, 'count'),
        **{k: pd.NamedAgg(column=v, aggfunc='sum')
           for k, v in {
               'إجمالي_الأقساط': COL_PREMIUM,
               'إجمالي_العمولات': 'العمولة',
           }.items()},
    ).reset_index()

    if 'الحالة' in df.columns:
        renewals = (
            df[df['الحالة'] == 'للتجديد']
            .groupby(COL_AGENT)
            .size()
            .reset_index(name='وثائق_للتجديد')
        )
        expired = (
            df[df['الحالة'] == 'منتهية']
            .groupby(COL_AGENT)
            .size()
            .reset_index(name='وثائق_منتهية')
        )
        counts = counts.merge(renewals, on=COL_AGENT, how='left')
        counts = counts.merge(expired, on=COL_AGENT, how='left')
        counts['وثائق_للتجديد'] = counts['وثائق_للتجديد'].fillna(0).astype(int)
        counts['وثائق_منتهية'] = counts['وثائق_منتهية'].fillna(0).astype(int)

    return counts.sort_values('إجمالي_الأقساط', ascending=False)


def calc_dues(df: pd.DataFrame) -> pd.DataFrame:
    """مستحقات كل وكيل (الرصيد غير المحصَّل)."""
    due = df[df[COL_BALANCE] > 0].copy() if COL_BALANCE in df.columns else df.copy()
    cols = [c for c in [COL_POLICY_NO, COL_CLIENT_NAME, COL_AGENT,
                        COL_PREMIUM, COL_COLLECTED, COL_BALANCE,
                        'العمولة', COL_EXPIRY_DATE, 'الحالة'] if c in due.columns]
    return due[cols].sort_values([COL_AGENT, COL_BALANCE], ascending=[True, False])


def calc_receivables_detail(df: pd.DataFrame) -> pd.DataFrame:
    """ذمم تفصيلية مرتّبة حسب الوكيل ثم تاريخ الانتهاء."""
    cols = [c for c in [COL_AGENT, COL_POLICY_NO, COL_CLIENT_NAME,
                        COL_POLICY_TYPE, COL_INSURER, COL_PREMIUM,
                        COL_BALANCE, COL_EXPIRY_DATE, 'الحالة'] if c in df.columns]
    return df[cols].sort_values(
        [COL_AGENT] + ([COL_EXPIRY_DATE] if COL_EXPIRY_DATE in df.columns else [])
    )


def build_summary(df: pd.DataFrame, commission_rate: float) -> list:
    """ملخص عام (قائمة أزواج مفتاح/قيمة)."""
    today = date.today()
    rows = [
        ('البيان', 'القيمة'),
        ('تاريخ التقرير', str(today)),
        ('نسبة العمولة المُطبَّقة', '{:.1%}'.format(commission_rate)),
        ('إجمالي الوثائق', len(df)),
        ('إجمالي الأقساط', round(df[COL_PREMIUM].sum(), 2)),
        ('إجمالي العمولات', round(df['العمولة'].sum(), 2)),
        ('إجمالي المحصَّل', round(df[COL_COLLECTED].sum(), 2)
         if COL_COLLECTED in df.columns else 'غير متاح'),
        ('إجمالي الذمم (الرصيد)', round(df[COL_BALANCE].sum(), 2)
         if COL_BALANCE in df.columns else 'غير متاح'),
        ('عدد الوكلاء', df[COL_AGENT].nunique()),
    ]
    if 'الحالة' in df.columns:
        rows += [
            ('وثائق للتجديد (خلال {} يومًا)'.format(RENEWAL_DAYS),
             int((df['الحالة'] == 'للتجديد').sum())),
            ('وثائق منتهية', int((df['الحالة'] == 'منتهية').sum())),
            ('وثائق نشطة', int((df['الحالة'] == 'نشطة').sum())),
        ]
    return rows


# ── تنسيق الكتابة في Excel ─────────────────────────────────────────────────────

def _header_font():
    return Font(bold=True, color=C_HEADER_FG, name='Calibri', size=11)


def _header_fill():
    return PatternFill(fill_type='solid',
                       start_color=C_HEADER_BG, end_color=C_HEADER_BG)


def _header_align():
    return Alignment(horizontal='center', vertical='center',
                     wrap_text=True, readingOrder=2)


def _border():
    s = Side(style='thin', color='BFBFBF')
    return Border(left=s, right=s, top=s, bottom=s)


def _cell_align():
    return Alignment(horizontal='right', vertical='center', readingOrder=2)


def write_df_sheet(ws, df: pd.DataFrame, title: str,
                   highlight_col: str = None, agent_sheet: bool = False) -> None:
    """كتابة DataFrame في ورقة عمل مع تنسيق كامل."""
    ws.title = title[:31]   # حد Excel لأسماء الأوراق
    ws.sheet_view.rightToLeft = True

    hf, hfill, ha = _header_font(), _header_fill(), _header_align()
    br = _border()

    # رأس الجدول
    for ci, col in enumerate(df.columns, 1):
        c = ws.cell(row=1, column=ci, value=col)
        c.font = hf; c.fill = hfill; c.alignment = ha; c.border = br
    ws.row_dimensions[1].height = 26
    ws.freeze_panes = 'A2'

    today = pd.Timestamp(date.today())

    for ri, row in enumerate(df.itertuples(index=False), 2):
        # تحديد لون الصف
        row_fill = None
        if highlight_col and highlight_col in df.columns:
            hi = list(df.columns).index(highlight_col)
            val = row[hi]
            try:
                exp = pd.Timestamp(val)
                days = (exp - today).days
                if days < 0:
                    row_fill = PatternFill(fill_type='solid',
                                          start_color=C_OVERDUE_BG, end_color=C_OVERDUE_BG)
                elif days <= RENEWAL_DAYS:
                    row_fill = PatternFill(fill_type='solid',
                                          start_color=C_SOON_BG, end_color=C_SOON_BG)
            except Exception:  # noqa: BLE001
                pass
        if row_fill is None and ri % 2 == 0:
            row_fill = PatternFill(fill_type='solid',
                                   start_color=C_ALT_ROW, end_color=C_ALT_ROW)

        for ci, val in enumerate(row, 1):
            # تحويل Timestamp لنص قابل للقراءة
            if isinstance(val, pd.Timestamp):
                val = val.strftime('%Y-%m-%d') if not pd.isnull(val) else ''
            c = ws.cell(row=ri, column=ci, value=val)
            c.border = br
            c.alignment = _cell_align()
            if row_fill:
                c.fill = row_fill

    # ضبط عرض الأعمدة
    for ci, col in enumerate(df.columns, 1):
        cl = get_column_letter(ci)
        col_vals = df.iloc[:, ci - 1].astype(str).str.len()
        max_len = max(len(str(col)), int(col_vals.max()) if len(col_vals) > 0 else 0)
        ws.column_dimensions[cl].width = min(max(max_len + 2, 12), 40)


def write_summary_sheet(ws, rows: list) -> None:
    ws.title = 'ملخص_عام'
    ws.sheet_view.rightToLeft = True
    hf, hfill, ha = _header_font(), _header_fill(), _header_align()
    br = _border()
    total_fill = PatternFill(fill_type='solid',
                             start_color=C_TOTAL_BG, end_color=C_TOTAL_BG)

    for ri, (label, value) in enumerate(rows, 1):
        c1 = ws.cell(row=ri, column=1, value=label)
        c2 = ws.cell(row=ri, column=2, value=value)
        for c in (c1, c2):
            c.border = br
            c.alignment = Alignment(horizontal='right', vertical='center', readingOrder=2)
        if ri == 1:
            for c in (c1, c2):
                c.font = hf; c.fill = hfill; c.alignment = ha
        elif str(label).startswith('إجمالي'):
            for c in (c1, c2):
                c.fill = total_fill; c.font = Font(bold=True, name='Calibri')

    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 22


# ── بناء التقرير الكامل ─────────────────────────────────────────────────────────

def build_report(xls_path: str, commission_rate: float, out_path: str) -> None:
    raw = load_file(xls_path)
    df = prepare(raw, commission_rate)

    wb = openpyxl.Workbook()

    # ١. ملخص عام
    ws_sum = wb.active
    write_summary_sheet(ws_sum, build_summary(df, commission_rate))

    # ٢. عمولات
    ws_comm = wb.create_sheet('عمولات')
    write_df_sheet(ws_comm, calc_commissions(df), 'عمولات',
                   highlight_col=COL_EXPIRY_DATE)

    # ٣. أداء الوكلاء
    ws_perf = wb.create_sheet('أداء_الوكلاء')
    write_df_sheet(ws_perf, calc_performance(df), 'أداء_الوكلاء')

    # ٤. مستحقات
    ws_dues = wb.create_sheet('مستحقات')
    write_df_sheet(ws_dues, calc_dues(df), 'مستحقات',
                   highlight_col=COL_EXPIRY_DATE)

    # ٥. ذمم تفصيلية
    ws_recv = wb.create_sheet('ذمم_تفصيل')
    write_df_sheet(ws_recv, calc_receivables_detail(df), 'ذمم_تفصيل',
                   highlight_col=COL_EXPIRY_DATE)

    # ٦. ورقة مستقلة لكل وكيل
    agents = sorted(df[COL_AGENT].dropna().unique())
    for agent in agents:
        agent_df = df[df[COL_AGENT] == agent].copy()
        # اسم الورقة: أول 31 حرف (حد Excel)
        sheet_name = str(agent)[:31]
        # تجنب تكرار الأسماء
        existing = [ws.title for ws in wb.worksheets]
        if sheet_name in existing:
            sheet_name = sheet_name[:28] + '_٢'
        ws_agent = wb.create_sheet(sheet_name)

        # أعمدة الورقة الفردية
        cols = [c for c in [COL_POLICY_NO, COL_CLIENT_NAME, COL_POLICY_TYPE,
                             COL_INSURER, COL_PREMIUM, COL_COLLECTED,
                             COL_BALANCE, 'العمولة', COL_EXPIRY_DATE, 'الحالة']
                if c in agent_df.columns]
        agent_out = agent_df[cols].sort_values(
            COL_EXPIRY_DATE if COL_EXPIRY_DATE in cols else cols[0]
        )
        write_df_sheet(ws_agent, agent_out, sheet_name,
                       highlight_col=COL_EXPIRY_DATE, agent_sheet=True)

        # صف إجمالي في أسفل الورقة
        last_row = ws_agent.max_row + 1
        total_fill = PatternFill(fill_type='solid',
                                 start_color=C_TOTAL_BG, end_color=C_TOTAL_BG)
        total_font = Font(bold=True, name='Calibri')
        br = _border()
        for ci in range(1, len(cols) + 1):
            c = ws_agent.cell(row=last_row, column=ci, value='')
            c.fill = total_fill; c.border = br

        # إجمالي القسط
        if COL_PREMIUM in cols:
            prem_ci = cols.index(COL_PREMIUM) + 1
            ws_agent.cell(row=last_row, column=prem_ci,
                          value=round(agent_df[COL_PREMIUM].sum(), 2)).font = total_font
        # إجمالي الرصيد
        if COL_BALANCE in cols:
            bal_ci = cols.index(COL_BALANCE) + 1
            ws_agent.cell(row=last_row, column=bal_ci,
                          value=round(agent_df[COL_BALANCE].sum(), 2)).font = total_font
        # إجمالي العمولة
        if 'العمولة' in cols:
            comm_ci = cols.index('العمولة') + 1
            ws_agent.cell(row=last_row, column=comm_ci,
                          value=round(agent_df['العمولة'].sum(), 2)).font = total_font
        # تسمية الإجمالي
        ws_agent.cell(row=last_row, column=1, value='الإجمالي').font = total_font

    wb.save(out_path)
    logger.info('تم حفظ التقرير في: %s', out_path)
    print('\nاكتمل ✔  التقرير محفوظ في: {}'.format(out_path))
    print('  الأوراق: ملخص_عام، عمولات، أداء_الوكلاء، مستحقات، ذمم_تفصيل + {} ورقة وكيل'.format(
        len(agents)))


# ── نقطة الدخول ───────────────────────────────────────────────────────────────

def main() -> None:
    parser = argparse.ArgumentParser(
        description='تقرير شامل لحسابات الوكلاء من ملف الذمم.',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    parser.add_argument('xls', help='مسار ملف الذمم (XLS أو XLSX)')
    parser.add_argument(
        '--commission',
        type=float,
        default=DEFAULT_COMMISSION_RATE,
        metavar='RATE',
        help='نسبة العمولة (0.0–1.0)، الافتراضي: {:.0%}'.format(DEFAULT_COMMISSION_RATE),
    )
    parser.add_argument(
        '--out',
        default='تقرير_الوكلاء.xlsx',
        help='مسار ملف الإخراج (الافتراضي: تقرير_الوكلاء.xlsx)',
    )
    args = parser.parse_args()

    if not os.path.isfile(args.xls):
        sys.exit('خطأ: الملف "{}" غير موجود.'.format(args.xls))
    if not (0.0 <= args.commission <= 1.0):
        sys.exit('خطأ: نسبة العمولة يجب أن تكون بين 0.0 و 1.0')

    build_report(args.xls, args.commission, args.out)


if __name__ == '__main__':
    main()
