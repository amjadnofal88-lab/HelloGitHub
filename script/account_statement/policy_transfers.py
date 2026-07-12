#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
policy_transfers.py — أمر حوالة لكل وثيقة بقيمة عمولتها

يقرأ كشف حساب الأهلية (PDF)، يحسب لكل بوليصة صافي إنتاجها
(إشعارات مدين − إشعارات دائن)، ثم يولّد أمر حوالة مستقل لكل بوليصة
بقيمة العمولة (25% التيرنتف / 27% نوفل) إلى آيبان المستفيد.

المخرجات:
  1) policy_transfers.xlsx — 3 شيتات: الملخص، أوامر الحوالات (واحد لكل
     وثيقة، بمرجع فريد)، والمستبعدات (بوالص صافيها صفر/سالب أو دون الحد).
  2) bank_bulk.csv — ملف حوالات مجمّعة بصيغة يقبلها قسم الحوالات
     (مرجع | مستفيد | آيبان | مبلغ | عملة | بيان) — UTF-8-BOM ليقرأه إكسل عربي.

⚠️ هذه أوامر/مستندات مطالبة — التنفيذ الفعلي يتم من طرف الأهلية أو البنك
   بعد اعتمادها. السكربت لا ينفّذ أي حركة مالية.

أمثلة:
    python policy_transfers.py كشف.pdf
    python policy_transfers.py كشف.pdf --rate 0.27
    python policy_transfers.py كشف.pdf --min 50        # ادمج ما دون 50₪ بحوالة واحدة
"""
import os
import re
import csv
import argparse
import datetime as dt
import collections
from decimal import Decimal, ROUND_HALF_UP

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from ahleia_statement import parse_statement, enrich

# ── إعدادات ──────────────────────────────────────────────────────────────
_BENEFICIARIES_CSV = os.path.join(os.path.dirname(__file__), 'beneficiaries.csv')
CURRENCY = 'ILS'


def load_beneficiary(beneficiary_id=1, csv_path=None):
    """Load a beneficiary record by BENEFICIARY_ID from beneficiaries.csv.

    Falls back to environment variables, then to placeholder strings if the
    CSV file is not found or the requested ID does not exist.
    """
    path = csv_path or _BENEFICIARIES_CSV
    if os.path.exists(path):
        with open(path, newline='', encoding='utf-8-sig') as fh:
            for row in csv.DictReader(fh):
                if int(row['BENEFICIARY_ID']) == beneficiary_id and row.get('STATUS', '').upper() == 'ACTIVE':
                    return {
                        'iban': row['IBAN'],
                        'name': row['BENEFICIARY_NAME'],
                        'currency': row['CURRENCY'] or CURRENCY,
                        'bank': row.get('BANK_NAME', ''),
                    }
    # Fallback: environment variables → placeholder
    return {
        'iban': os.environ.get('BENEFICIARY_IBAN', 'CONFIGURE_BENEFICIARY_IBAN'),
        'name': os.environ.get('BENEFICIARY_NAME', 'CONFIGURE_BENEFICIARY_NAME'),
        'currency': CURRENCY,
        'bank': '',
    }
RATES = {'475151': Decimal('0.27'), '506322': Decimal('0.25')}
DEFAULT_RATE = Decimal('0.25')
Q = Decimal('0.01')

HDR_FILL = PatternFill(fill_type='solid', fgColor='1F4E78')
HDR_FONT = Font(bold=True, color='FFFFFF', name='Arial', size=11)
BODY = Font(name='Arial', size=10)
MONEY = '#,##0.00;[Red]-#,##0.00;-'


def money(x):
    return Decimal(str(x)).quantize(Q, rounding=ROUND_HALF_UP)


def per_policy(rows):
    """صافي إنتاج كل بوليصة + تاريخ أول إشعار مدين لها."""
    pol = collections.defaultdict(lambda: {'dn': Decimal(0), 'cn': Decimal(0),
                                           'first': None, 'entries': 0})
    for r in rows:
        p = r['policy_no']
        if not p:
            continue
        if r['type'] == 'DebitNote':
            pol[p]['dn'] += Decimal(str(r['debit']))
            pol[p]['entries'] += 1
            if pol[p]['first'] is None:
                pol[p]['first'] = r['date']
        elif r['type'] == 'CreditNote':
            pol[p]['cn'] += Decimal(str(r['credit']))
    return pol


def build_orders(pdf_path, rate=None, min_amount=Decimal('0'), beneficiary=None):
    info, rows = parse_statement(pdf_path)
    rows = enrich(rows)
    acct = re.sub(r'\D', '', info.get('account', ''))[-6:]
    rate = rate or RATES.get(acct, DEFAULT_RATE)
    ben = beneficiary or load_beneficiary()

    pol = per_policy(rows)
    today = dt.date.today().strftime('%Y%m%d')

    orders, excluded = [], []
    seq = 0
    misc_total = Decimal(0)     # ما دون الحد الأدنى — يُجمع بحوالة واحدة
    misc_policies = []

    for pno in sorted(pol):
        d = pol[pno]
        net = d['dn'] - d['cn']
        comm = money(net * rate)
        if net <= 0:
            excluded.append([pno, float(d['dn']), float(d['cn']), float(net),
                             'صافي صفر أو سالب (ملغاة/معكوسة)'])
            continue
        if comm < min_amount:
            misc_total += comm
            misc_policies.append(pno)
            excluded.append([pno, float(d['dn']), float(d['cn']), float(net),
                             f'دون الحد {min_amount}₪ — ضُمّت للحوالة المجمّعة'])
            continue
        seq += 1
        ref = f'COMM-{acct}-{today}-{seq:04d}'
        orders.append({
            'ref': ref, 'policy': pno, 'first_date': d['first'],
            'net': net, 'rate': rate, 'amount': comm,
            'note': f'عمولة {rate*100:.0f}% عن الوثيقة {pno}',
        })

    if misc_total > 0:
        seq += 1
        ref = f'COMM-{acct}-{today}-{seq:04d}'
        orders.append({
            'ref': ref, 'policy': f'مجمّعة ({len(misc_policies)} وثيقة)',
            'first_date': '', 'net': Decimal(0), 'rate': rate,
            'amount': money(misc_total),
            'note': f'عمولة مجمّعة عن {len(misc_policies)} وثيقة دون الحد الأدنى',
        })

    return info, rate, orders, excluded, ben


def export(info, rate, orders, excluded, out_xlsx, out_csv, beneficiary=None):
    ben = beneficiary or load_beneficiary()
    iban = ben['iban']
    beneficiary_name = ben['name']
    currency = ben['currency']
    total = sum(o['amount'] for o in orders)
    n = len(orders)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'الملخص'
    ws.sheet_view.rightToLeft = True
    ws.column_dimensions['A'].width = 32
    ws.column_dimensions['B'].width = 24
    ws['A1'] = 'أوامر حوالات العمولة — وثيقة بوثيقة'
    ws['A1'].font = Font(bold=True, size=14, name='Arial', color='1F4E78')
    facts = [
        ('رقم الحساب', info.get('account', '')),
        ('الفترة', f"{info.get('from','')} → {info.get('to','')}"),
        ('نسبة العمولة', f'{rate*100:.0f}%'),
        ('عدد أوامر الحوالات', n),
        ('إجمالي الحوالات', float(total)),
        ('المستفيد', beneficiary_name),
        ('الآيبان', iban),
        ('العملة', currency),
        ('تاريخ الإصدار', dt.date.today().strftime('%d-%m-%Y')),
    ]
    for i, (k, v) in enumerate(facts, start=3):
        ws.cell(i, 1, k).font = Font(bold=True, name='Arial')
        c = ws.cell(i, 2, v)
        c.font = BODY
        if isinstance(v, float):
            c.number_format = MONEY
        if k == 'الآيبان':
            c.number_format = '@'

    def sheet(title, headers, data, widths, money_cols=()):
        w = wb.create_sheet(title)
        w.sheet_view.rightToLeft = True
        w.append(headers)
        for c in w[1]:
            c.font, c.fill = HDR_FONT, HDR_FILL
            c.alignment = Alignment(horizontal='center')
        for r in data:
            w.append(r)
        for i, wd in enumerate(widths, 1):
            w.column_dimensions[get_column_letter(i)].width = wd
        for col in money_cols:
            for c in w[get_column_letter(col)][1:]:
                c.number_format = MONEY
        for row in w.iter_rows(min_row=2):
            for c in row:
                c.font = BODY
        w.freeze_panes = 'A2'
        w.auto_filter.ref = w.dimensions
        return w

    ord_rows = [[o['ref'], o['policy'], o['first_date'], float(o['net']),
                 f"{o['rate']*100:.0f}%", float(o['amount']), iban, o['note']]
                for o in orders]
    w2 = sheet('أوامر الحوالات',
               ['المرجع', 'رقم الوثيقة', 'تاريخ الإصدار', 'صافي الإنتاج',
                'النسبة', 'مبلغ الحوالة', 'آيبان المستفيد', 'البيان'],
               ord_rows, [24, 24, 13, 14, 8, 14, 32, 42], money_cols=(4, 6))
    w2.append(['الإجمالي', '', '', '', '', f'=SUM(F2:F{n+1})', '', ''])
    w2.cell(n + 2, 1).font = Font(bold=True, name='Arial')
    w2.cell(n + 2, 6).font = Font(bold=True, name='Arial')
    w2.cell(n + 2, 6).number_format = MONEY
    for c in w2['G'][1:]:
        c.number_format = '@'

    sheet('المستبعدات',
          ['رقم الوثيقة', 'مدين', 'دائن', 'الصافي', 'السبب'],
          excluded, [24, 13, 13, 13, 40], money_cols=(2, 3, 4))

    wb.save(out_xlsx)

    # ملف الحوالات المجمّعة للبنك
    with open(out_csv, 'w', newline='', encoding='utf-8-sig') as fh:
        wr = csv.writer(fh)
        wr.writerow(['REFERENCE', 'BENEFICIARY_NAME', 'BENEFICIARY_IBAN',
                     'AMOUNT', 'CURRENCY', 'DETAILS'])
        for o in orders:
            wr.writerow([o['ref'], beneficiary_name, iban,
                         f"{o['amount']:.2f}", currency, o['note']])
    return total


def main():
    ap = argparse.ArgumentParser(description='أمر حوالة لكل وثيقة بقيمة عمولتها')
    ap.add_argument('pdf', help='كشف حساب الأهلية (PDF)')
    ap.add_argument('--rate', type=float, help='تجاوز النسبة (مثال 0.27)')
    ap.add_argument('--min', type=float, default=0,
                    help='حد أدنى للحوالة — ما دونه يُجمع بحوالة واحدة')
    ap.add_argument('--beneficiary', type=int, default=1,
                    help='رقم المستفيد في beneficiaries.csv (افتراضي: 1)')
    ap.add_argument('-o', '--output', default='policy_transfers.xlsx')
    ap.add_argument('--csv', default='bank_bulk.csv')
    a = ap.parse_args()

    rate = Decimal(str(a.rate)) if a.rate else None
    ben = load_beneficiary(a.beneficiary)
    info, rate, orders, excluded, ben = build_orders(a.pdf, rate, money(a.min), ben)
    total = export(info, rate, orders, excluded, a.output, a.csv, ben)

    print(f"الحساب {info.get('account','')} | نسبة {rate*100:.0f}%")
    print(f"المستفيد: {ben['name']} | آيبان: {ben['iban']}")
    print(f"أوامر حوالات: {len(orders)} | الإجمالي: {total:,.2f}₪")
    print(f"مستبعدات: {len(excluded)}")
    print(f"Saved: {a.output}  |  {a.csv}")


if __name__ == '__main__':
    main()
