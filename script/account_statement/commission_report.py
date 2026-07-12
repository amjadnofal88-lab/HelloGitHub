#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
commission_report.py — كشف العمولة المحتجزة وخطة الدفعات

يقرأ كشف حساب الأهلية (PDF) ويحسب العمولة المحتجزة بناءً على نسبة تعاقدية،
ثم يولّد خطة دفعات ويصدرها إلى مصنف إكسل عربي RTL.

النسب التعاقدية تُقرأ من متغير البيئة COMMISSION_RATES (JSON) أو تُمرَّر
بـ --rate. لا يوجد أي بيانك مالية (IBAN / بنك / مستفيد) مضمّنة في الكود —
جميعها تأتي من متغيرات البيئة أو سطر الأوامر.

متغيرات البيئة المدعومة:
    COMMISSION_RATES   JSON لنسب الحسابات  مثال: '{"475151": 0.27, "506322": 0.25}'
    BENEFICIARY_NAME   اسم المستفيد للبيان (اختياري)
    BENEFICIARY_BANK   اسم البنك للبيان (اختياري)

أمثلة:
    # حساب وخطة — 12 دفعة شهرية
    python commission_report.py statement.pdf --installments 12

    # نسبة مخصصة + مبلغ ثابت لكل دفعة
    python commission_report.py statement.pdf --rate 0.27 --amount-per 5000

    # مع تحديد ملف الإخراج
    python commission_report.py statement.pdf --installments 6 -o comm_plan.xlsx
"""
import os
import re
import json
import argparse
import datetime as dt
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from ahleia_statement import parse_statement, enrich

# ── إعدادات افتراضية ──────────────────────────────────────────────────────
_DEFAULT_RATES = {
    '475151': Decimal('0.27'),
    '506322': Decimal('0.25'),
}
DEFAULT_RATE = Decimal('0.25')
Q = Decimal('0.01')

# قراءة النسب من متغير البيئة إن وُجد
_env_rates = os.environ.get('COMMISSION_RATES', '')
if _env_rates:
    try:
        _DEFAULT_RATES = {k: Decimal(str(v)) for k, v in json.loads(_env_rates).items()}
    except (json.JSONDecodeError, Exception):
        pass

# بيانات المستفيد — من متغيرات البيئة فقط، بدون قيم افتراضية مشفّرة
BENEFICIARY_NAME = os.environ.get('BENEFICIARY_NAME', '')
BENEFICIARY_BANK = os.environ.get('BENEFICIARY_BANK', '')

# ── أنماط الإكسل ──────────────────────────────────────────────────────────
HDR_FILL = PatternFill(fill_type='solid', fgColor='1F4E78')
HDR_FONT = Font(bold=True, color='FFFFFF', name='Arial', size=11)
BODY_FONT = Font(name='Arial', size=10)
BOLD_FONT = Font(bold=True, name='Arial', size=11)
RED_BOLD = Font(bold=True, name='Arial', size=11, color='C00000')
MONEY_FMT = '#,##0.00;[Red]-#,##0.00;-'
TITLE_FONT = Font(bold=True, size=14, name='Arial', color='1F4E78')


def _money(value):
    return Decimal(str(value)).quantize(Q, rounding=ROUND_HALF_UP)


# ── 1) الحساب ────────────────────────────────────────────────────────────
def compute(pdf_path, rate=None):
    """يحسب العمولة المحتجزة من ملف PDF.

    :param pdf_path: مسار ملف PDF لكشف الحساب.
    :param rate: نسبة العمولة (Decimal). إذا لم تُحدَّد تُستخرج من COMMISSION_RATES.
    :return: dict يحتوي على نتائج الحساب.
    """
    info, rows = parse_statement(pdf_path)
    rows = enrich(rows)
    if not rows:
        raise SystemExit('خطأ: لم يُستخرج أي سطر من الكشف. تحقق من الملف.')

    acct = re.sub(r'\D', '', info.get('account', ''))[-6:]
    effective_rate = rate if rate is not None else _DEFAULT_RATES.get(acct, DEFAULT_RATE)

    production = sum(
        Decimal(str(r['debit'])) for r in rows if r['type'] == 'DebitNote'
    )
    comm_cr = sum(
        Decimal(str(r['credit'])) for r in rows
        if 'عمول' in r.get('notes', '')
    )
    comm_dr = sum(
        Decimal(str(r['debit'])) for r in rows
        if 'عمول' in r.get('notes', '')
    )
    credited = comm_cr - comm_dr
    due = production * effective_rate
    withheld = due - credited

    return {
        'account': info.get('account', ''),
        'supplier': info.get('supplier', ''),
        'period': '{} ← {}'.format(info.get('to', ''), info.get('from', '')),
        'rate': effective_rate,
        'production': _money(production),
        'commission_due': _money(due),
        'commission_credited': _money(credited),
        'withheld': _money(withheld),
        'final_balance': _money(Decimal(str(rows[-1]['balance']))),
        'row_count': len(rows),
    }


# ── 2) خطة الدفعات ───────────────────────────────────────────────────────
def make_plan(total, n=None, amount_per=None, every=30, start=None):
    """يبني قائمة دفعات.

    :param total: إجمالي العمولة المحتجزة (Decimal).
    :param n: عدد الدفعات (يُستخدم إذا لم يُحدَّد amount_per).
    :param amount_per: مبلغ ثابت لكل دفعة.
    :param every: عدد الأيام بين الدفعات (افتراضي 30).
    :param start: تاريخ أول دفعة (datetime.date). افتراضياً اليوم + every.
    :return: list من dict {no, amount, due}.
    """
    total = _money(total)
    if total <= 0:
        raise SystemExit('لا توجد عمولة محتجزة موجبة — لا حاجة لخطة دفعات.')

    start = start or (dt.date.today() + dt.timedelta(days=every))

    if amount_per is not None:
        per = _money(amount_per)
        if per <= 0:
            raise SystemExit('--amount-per يجب أن يكون موجباً.')
        n = max(1, int((total / per).to_integral_value(rounding=ROUND_HALF_UP)))
    elif n:
        if n <= 0:
            raise SystemExit('--installments يجب أن يكون موجباً.')
        per = _money(total / Decimal(n))
    else:
        raise SystemExit('حدّد --installments أو --amount-per.')

    plan = []
    remaining = total
    for i in range(1, n + 1):
        amt = per if i < n else _money(remaining)
        amt = min(amt, remaining)
        if amt <= 0:
            break
        due_date = start + dt.timedelta(days=every * (i - 1))
        plan.append({'no': i, 'amount': amt, 'due': due_date})
        remaining = _money(remaining - amt)
        if remaining <= 0:
            break

    return plan


def _note_for(inst, calc):
    """بيان الدفعة للإكسل."""
    parts = [
        'عمولة مستحقة — دفعة {}/{}'.format(inst['no'], inst['no']),
        'حساب {}'.format(calc['account']),
        'نسبة {:.0f}%'.format(calc['rate'] * 100),
    ]
    if BENEFICIARY_NAME:
        parts.append('المستفيد: {}'.format(BENEFICIARY_NAME))
    return ' | '.join(parts)


# ── 3) تصدير إكسل ────────────────────────────────────────────────────────
def export_plan(calc, plan, out='commission_plan.xlsx'):
    """يصدر خطة الدفعات إلى مصنف إكسل.

    :param calc: نتائج دالة compute().
    :param plan: قائمة الدفعات من make_plan().
    :param out: مسار ملف الإخراج.
    :return: مسار الملف المُنشأ.
    """
    wb = openpyxl.Workbook()

    # ── ورقة الاحتساب ────────────────────────────────────────────────────
    ws = wb.active
    ws.title = 'الاحتساب'
    ws.sheet_view.rightToLeft = True
    ws.column_dimensions['A'].width = 34
    ws.column_dimensions['B'].width = 24

    ws['A1'] = 'كشف العمولة المحتجزة'
    ws['A1'].font = TITLE_FONT
    ws['A2'] = 'تاريخ التقرير: {}'.format(dt.date.today().strftime('%d-%m-%Y'))
    ws['A2'].font = Font(italic=True, size=9, name='Arial')

    facts = [
        ('رقم الحساب',           calc['account'],                  None,     BOLD_FONT),
        ('المورّد',              calc['supplier'],                  None,     BODY_FONT),
        ('الفترة',               calc['period'],                    None,     BODY_FONT),
        ('عدد القيود',           calc['row_count'],                 None,     BODY_FONT),
        ('نسبة العمولة التعاقدية', '{:.0f}%'.format(calc['rate'] * 100), None, BOLD_FONT),
        ('',                     '',                                None,     BODY_FONT),
        ('إجمالي الإنتاج',       float(calc['production']),         MONEY_FMT, BODY_FONT),
        ('العمولة المستحقة',      float(calc['commission_due']),     MONEY_FMT, BODY_FONT),
        ('العمولة المقيّدة بالكشف', float(calc['commission_credited']), MONEY_FMT, BODY_FONT),
        ('العمولة المحتجزة',      float(calc['withheld']),           MONEY_FMT, RED_BOLD),
        ('الرصيد الختامي بالكشف', float(calc['final_balance']),      MONEY_FMT, BODY_FONT),
    ]

    if BENEFICIARY_NAME:
        facts.append(('المستفيد', BENEFICIARY_NAME, None, BODY_FONT))
    if BENEFICIARY_BANK:
        facts.append(('البنك', BENEFICIARY_BANK, None, BODY_FONT))

    for row_idx, (label, value, fmt, font) in enumerate(facts, start=4):
        lbl_cell = ws.cell(row=row_idx, column=1, value=label)
        lbl_cell.font = font
        val_cell = ws.cell(row=row_idx, column=2, value=value)
        val_cell.font = font
        if fmt:
            val_cell.number_format = fmt

    # ── ورقة الدفعات ─────────────────────────────────────────────────────
    ws2 = wb.create_sheet('الدفعات')
    ws2.sheet_view.rightToLeft = True

    headers = ['رقم الدفعة', 'المبلغ (₪)', 'تاريخ الاستحقاق', 'البيان']
    ws2.append(headers)
    for cell in ws2[1]:
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = Alignment(horizontal='center')

    for p in plan:
        ws2.append([
            p['no'],
            float(p['amount']),
            p['due'].strftime('%d-%m-%Y'),
            _note_for(p, calc),
        ])
        ws2.cell(ws2.max_row, 2).number_format = MONEY_FMT
        ws2.cell(ws2.max_row, 2).font = BODY_FONT

    # صف الإجمالي
    n = len(plan)
    ws2.append(['الإجمالي', float(sum(p['amount'] for p in plan)), '', ''])
    total_row = ws2.max_row
    ws2.cell(total_row, 1).font = BOLD_FONT
    ws2.cell(total_row, 2).font = BOLD_FONT
    ws2.cell(total_row, 2).number_format = MONEY_FMT
    ws2.cell(total_row, 2).fill = PatternFill(fill_type='solid', fgColor='D9E1F2')

    for col, width in zip('ABCD', [12, 18, 16, 80]):
        ws2.column_dimensions[col].width = width
    ws2.freeze_panes = 'A2'
    ws2.auto_filter.ref = 'A1:D{}'.format(n + 1)

    wb.save(out)
    return out


# ── main ──────────────────────────────────────────────────────────────────
def main():
    ap = argparse.ArgumentParser(
        description='يحسب العمولة المحتجزة ويبني خطة دفعات من كشف حساب PDF.'
    )
    ap.add_argument('pdf', help='مسار ملف PDF لكشف الحساب')
    ap.add_argument('--rate', type=float, default=None,
                    help='نسبة العمولة التعاقدية (مثال: 0.27). '
                         'إذا لم تُحدَّد تُستخرج من COMMISSION_RATES أو الافتراضي 25%%.')
    ap.add_argument('--installments', type=int, default=None,
                    help='عدد الدفعات الشهرية')
    ap.add_argument('--amount-per', type=float, default=None,
                    help='مبلغ ثابت لكل دفعة (بالشيكل)')
    ap.add_argument('--every', type=int, default=30,
                    help='عدد الأيام بين الدفعات (افتراضي: 30)')
    ap.add_argument('--start', default=None,
                    help='تاريخ أول دفعة بصيغة YYYY-MM-DD (افتراضي: اليوم + --every)')
    ap.add_argument('-o', '--output', default=None,
                    help='مسار ملف الإخراج .xlsx (افتراضي: commission_plan_YYYYMMDD.xlsx)')
    a = ap.parse_args()

    if a.installments is None and a.amount_per is None:
        ap.error('حدّد --installments أو --amount-per.')

    rate = Decimal(str(a.rate)) if a.rate is not None else None
    calc = compute(a.pdf, rate)

    # ── طباعة الملخص ──────────────────────────────────────────────────────
    print()
    print('═' * 44)
    print('  كشف العمولة المحتجزة')
    print('═' * 44)
    print('  الحساب         : {}'.format(calc['account']))
    print('  المورّد        : {}'.format(calc['supplier']))
    print('  الفترة         : {}'.format(calc['period']))
    print('  نسبة العمولة   : {:.0f}%'.format(calc['rate'] * 100))
    print('─' * 44)
    print('  إجمالي الإنتاج  : {:>14,.2f} ₪'.format(calc['production']))
    print('  العمولة المستحقة: {:>14,.2f} ₪'.format(calc['commission_due']))
    print('  المقيّدة منها  : {:>14,.2f} ₪'.format(calc['commission_credited']))
    print('  المحتجزة       : {:>14,.2f} ₪  ◄'.format(calc['withheld']))
    print('═' * 44)

    start = dt.date.fromisoformat(a.start) if a.start else None
    plan = make_plan(calc['withheld'], a.installments, a.amount_per, a.every, start)

    print()
    print('  خطة الدفعات ({} دفعة):'.format(len(plan)))
    print('  {:<8} {:>14}   {}'.format('رقم', 'المبلغ ₪', 'الاستحقاق'))
    print('  ' + '─' * 36)
    for p in plan:
        print('  دفعة {:>2}   {:>14,.2f}   {}'.format(
            p['no'], p['amount'], p['due'].strftime('%d-%m-%Y')))
    total = sum(p['amount'] for p in plan)
    print('  ' + '─' * 36)
    print('  الإجمالي  {:>14,.2f} ₪'.format(total))
    print()

    out = a.output or 'commission_plan_{}.xlsx'.format(dt.date.today().strftime('%Y%m%d'))
    export_plan(calc, plan, out)
    print('تم الحفظ: {}'.format(out))


if __name__ == '__main__':
    main()
