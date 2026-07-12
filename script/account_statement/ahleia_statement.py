#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
ahleia_statement.py — كشف حساب الأهلية إلى إكسل
يقرأ كشف حساب PDF (AHLEIA / ESKA format) + ملف ذمم (Crystal .xls/.xlsx)
ويولّد مصنف إكسل: ملخص، الحركات، الشيكات، الشيكات المرتجعة، الذمم، مطابقة البوالص.

Usage:
    python ahleia_statement.py statement.pdf --receivables receivables.xlsx -o output.xlsx
    python ahleia_statement.py statement.pdf --icloud --icloud-folder "Ahleia Statements"
"""
import os
import re
import shutil
import argparse
import subprocess
from pathlib import Path

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

NUM = r'-?[\d,]+\.\d{2}'
ROW_RE = re.compile(
    r'^\s*(\d{2}-\d{2}-\d{4})\s+(\S+)\s+([A-Za-z]+)\s+(.*?)\s*'
    rf'({NUM})\s+([A-Z]{{3}})\s+({NUM})\s+({NUM})\s+({NUM})\s*$'
)
OB_RE = re.compile(rf'^\s*(\d{{2}}-\d{{2}}-\d{{4}})\s+OB\s+(\d{{2}}-\d{{2}}-\d{{4}})\s+({NUM})\s+({NUM})\s*$')
DUE_RE = re.compile(r'\b(\d{2}-\d{2}-\d{4})\b')
BIDI_RE = re.compile('[\u200e\u200f\u202a-\u202e\u2066-\u2069\ufeff]')
CHEQUE_RE = re.compile(r'Cheque\s*(?:No[:.]?\s*)?(\d{4,})', re.IGNORECASE)
POLICY_RE = re.compile(r'(\d{2}-\d-\d{2}-\d{2}-\d{6})')
JTYPE_RE = re.compile(r'^(DebitNote|CreditNote|ReceiptVoucher|PaymentVoucher|JournalVoucher|ReturnedCheque)')


def f(s):
    return float(s.replace(',', ''))


def extract_pdf_text(pdf_path):
    out = subprocess.run(['pdftotext', '-layout', str(pdf_path), '-'],
                         capture_output=True, text=True, check=True)
    return out.stdout


def parse_statement(pdf_path):
    """Parse the statement PDF into (header_info, rows)."""
    text = extract_pdf_text(pdf_path)
    lines = text.splitlines()

    info = {}
    for ln in lines[:40]:
        m = re.search(r'Account No\s*:\s*(\S+)', ln)
        if m:
            info['account'] = m.group(1)
        m = re.search(r'From Date\s*:\s*(\S+)\s+To Date\s*:\s*(\S+)', ln)
        if m:
            info['from'], info['to'] = m.group(1), m.group(2)
        m = re.search(r'Supplier\s*:\s*(.+)', ln)
        if m and 'supplier' not in info:
            info['supplier'] = m.group(1).strip()

    rows = []
    current = None
    for raw in lines:
        ln = BIDI_RE.sub('', raw)
        s = ln.rstrip()
        if not s.strip():
            continue
        if re.search(r'Journal Date\s+Journal No|Page \d+ of \d+|Customer Total|Period Balance|Statement of Account|Supplier\s*:|From Date|Account No|Business Line|Address|User:|Date:', s):
            continue

        m = ROW_RE.match(s)
        if m:
            date, jno, jtype, middle, amount, cur, debit, credit, bal = m.groups()
            # due date may sit before or after the notes text (RTL layout)
            dm = DUE_RE.search(middle)
            if dm:
                due = dm.group(1)
                notes = (middle[:dm.start()] + ' ' + middle[dm.end():]).strip()
            else:
                due, notes = date, middle.strip()
            current = {
                'date': date, 'journal_no': jno, 'type': jtype, 'due': due,
                'notes': notes, 'amount': f(amount), 'currency': cur,
                'debit': f(debit), 'credit': f(credit), 'balance': f(bal),
            }
            rows.append(current)
            continue

        m = OB_RE.match(s)
        if m:
            current = {'date': m.group(1), 'journal_no': '', 'type': 'OB',
                       'due': m.group(2), 'notes': 'Opening Balance',
                       'amount': f(m.group(3)), 'currency': '',
                       'debit': 0.0, 'credit': 0.0, 'balance': f(m.group(4))}
            rows.append(current)
            continue

        # continuation line -> append to previous notes
        if current is not None and re.match(r'^\s{10,}\S', ln):
            current['notes'] = (current['notes'] + ' ' + s.strip()).strip()

    # fix rows where type wasn't detected: layout puts type in its own column,
    # captured inside `jno...notes`; re-scan
    for r in rows:
        if not r['type']:
            m = JTYPE_RE.search(r['journal_no'] + ' ' + r['notes'])
            r['type'] = m.group(1) if m else '?'
    return info, rows


def enrich(rows):
    for r in rows:
        notes_clean = r['notes'].replace(' ', '')
        if 'ReturnedCheque' in notes_clean:
            r['type'] = 'ReturnedCheque'
        elif 'CanceledCheque' in notes_clean:
            r['type'] = 'CanceledCheque'
        m = CHEQUE_RE.search(notes_clean)
        r['cheque_no'] = m.group(1) if m else ''
        m = POLICY_RE.search(notes_clean)
        r['policy_no'] = m.group(1) if m else ''
    return rows


HDR_FILL = PatternFill(fill_type='solid', fgColor='1F4E78')
HDR_FONT = Font(bold=True, color='FFFFFF', name='Arial', size=11)
BODY_FONT = Font(name='Arial', size=10)
RED_FONT = Font(name='Arial', size=10, color='CC0000', bold=True)
MONEY_FMT = '#,##0.00;[Red]-#,##0.00;-'


def write_sheet(wb, title, headers, data_rows, widths, rtl=True, money_cols=()):
    ws = wb.create_sheet(title)
    ws.sheet_view.rightToLeft = rtl
    ws.append(headers)
    for c in ws[1]:
        c.font, c.fill = HDR_FONT, HDR_FILL
        c.alignment = Alignment(horizontal='center')
    for row in data_rows:
        ws.append(row)
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for col in money_cols:
        for cell in ws[get_column_letter(col)][1:]:
            cell.number_format = MONEY_FMT
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions
    return ws


def build_workbook(info, rows, receivables_df, out_path):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    n = len(rows)
    final_balance = rows[-1]['balance'] if rows else 0
    tx_headers = ['التاريخ', 'رقم القيد', 'نوع الحركة', 'تاريخ الاستحقاق',
                  'البيان', 'المبلغ', 'العملة', 'مدين', 'دائن', 'الرصيد',
                  'رقم الشيك', 'رقم البوليصة']
    tx_rows = [[r['date'], r['journal_no'], r['type'], r['due'], r['notes'],
                r['amount'], r['currency'], r['debit'], r['credit'],
                r['balance'], r['cheque_no'], r['policy_no']] for r in rows]
    write_sheet(wb, 'الحركات', tx_headers, tx_rows,
                [12, 30, 16, 14, 55, 12, 8, 13, 13, 14, 12, 20],
                money_cols=(6, 8, 9, 10))

    cheq = [r for r in rows if r['cheque_no']]
    ch_rows = [[r['cheque_no'], r['date'], r['due'], r['type'], r['amount'],
                r['currency'], r['notes'][:80]] for r in cheq]
    write_sheet(wb, 'الشيكات', ['رقم الشيك', 'تاريخ القيد', 'الاستحقاق',
                                'نوع الحركة', 'المبلغ', 'العملة', 'البيان'],
                ch_rows, [14, 12, 12, 16, 13, 8, 60], money_cols=(5,))

    ret = [r for r in rows if r['type'] == 'ReturnedCheque']
    ret_rows = [[r['cheque_no'], r['date'], r['due'], r['amount'],
                 r['currency'], r['debit'], r['notes'][:80]] for r in ret]
    ws_ret = write_sheet(wb, 'الشيكات المرتجعة',
                         ['رقم الشيك', 'تاريخ الإرجاع', 'الاستحقاق', 'المبلغ',
                          'العملة', 'مدين', 'البيان'],
                         ret_rows, [14, 12, 12, 13, 8, 13, 60], money_cols=(4, 6))
    for row in ws_ret.iter_rows(min_row=2):
        for c in row:
            c.font = RED_FONT

    # receivables sheet
    if receivables_df is not None:
        rec_cols = ['BENEFICIARY', 'DOCUMENT_NO', 'DOCUMENT_TYPE', 'CLASS',
                    'ISSUE_DATE', 'NET_PREMIUM_LC', 'CASH', 'CHEQUES', 'NETTING',
                    'TOTAL_COLLECTION', 'AR_BALANCE', 'RETURNED_CHEQUES',
                    'NET_AR_BALANCE']
        df = receivables_df[rec_cols].copy()
        df['ISSUE_DATE'] = pd.to_datetime(df['ISSUE_DATE']).dt.strftime('%d-%m-%Y')
        rec_rows = df.values.tolist()
        write_sheet(wb, 'الذمم',
                    ['المؤمَّن له', 'رقم البوليصة', 'نوع الوثيقة', 'الفرع',
                     'تاريخ الإصدار', 'صافي القسط', 'نقداً', 'شيكات', 'تسوية',
                     'إجمالي التحصيل', 'رصيد الذمة', 'شيكات مرتجعة',
                     'صافي رصيد الذمة'],
                    rec_rows, [28, 22, 16, 12, 13, 13, 11, 12, 11, 14, 13, 13, 14],
                    money_cols=(6, 7, 8, 9, 10, 11, 12, 13))

        # reconciliation: policies with outstanding balance vs statement entries
        stmt_policies = {}
        for r in rows:
            if r['policy_no']:
                stmt_policies.setdefault(r['policy_no'], []).append(r)
        rec_open = receivables_df[receivables_df['NET_AR_BALANCE'] > 0]
        match_rows = []
        for _, rr in rec_open.iterrows():
            pno = str(rr['DOCUMENT_NO']).split('-R-')[0].split('-E-')[0]
            hits = stmt_policies.get(pno, [])
            match_rows.append([
                rr['BENEFICIARY'], rr['DOCUMENT_NO'],
                float(rr['NET_PREMIUM_LC']), float(rr['NET_AR_BALANCE']),
                float(rr['RETURNED_CHEQUES']), len(hits),
                sum(h['debit'] for h in hits), sum(h['credit'] for h in hits),
            ])
        write_sheet(wb, 'مطابقة البوالص',
                    ['المؤمَّن له', 'رقم البوليصة', 'صافي القسط',
                     'صافي رصيد الذمة', 'شيكات مرتجعة',
                     'عدد قيود الكشف', 'مدين بالكشف', 'دائن بالكشف'],
                    match_rows, [28, 22, 13, 14, 13, 12, 13, 13],
                    money_cols=(3, 4, 5, 7, 8))

    # ── Summary sheet (formulas, placed first) ──
    ws = wb.create_sheet('الملخص', 0)
    ws.sheet_view.rightToLeft = True
    ws.column_dimensions['A'].width = 34
    ws.column_dimensions['B'].width = 24
    title = Font(bold=True, size=14, name='Arial', color='1F4E78')
    lbl = Font(bold=True, name='Arial', size=11)

    def put(r, a, b, fmt=None, bold=False):
        ws.cell(row=r, column=1, value=a).font = lbl if bold else BODY_FONT
        c = ws.cell(row=r, column=2, value=b)
        c.font = lbl if bold else BODY_FONT
        if fmt:
            c.number_format = fmt

    ws['A1'] = 'كشف حساب الأهلية — ملخص'
    ws['A1'].font = title
    put(3, 'رقم الحساب', info.get('account', ''), bold=True)
    put(4, 'المورّد', info.get('supplier', ''))
    put(5, 'الفترة', "{} \u2192 {}".format(info.get('from', ''), info.get('to', '')))
    put(6, 'عدد القيود', n)
    put(8, 'إجمالي المدين', '=SUM(الحركات!H2:H{})'.format(n + 1), MONEY_FMT, bold=True)
    put(9, 'إجمالي الدائن', '=SUM(الحركات!I2:I{})'.format(n + 1), MONEY_FMT, bold=True)
    put(10, 'الرصيد (مدين - دائن)', '=B8-B9', MONEY_FMT, bold=True)
    put(11, 'الرصيد الختامي بالكشف', final_balance, MONEY_FMT, bold=True)
    put(13, 'عدد إشعارات المدين (بوالص)', '=COUNTIF(الحركات!C2:C{},"DebitNote")'.format(n + 1))
    put(14, 'عدد إشعارات الدائن', '=COUNTIF(الحركات!C2:C{},"CreditNote")'.format(n + 1))
    put(15, 'عدد سندات القبض', '=COUNTIF(الحركات!C2:C{},"ReceiptVoucher")'.format(n + 1))
    put(16, 'عدد قيود التسوية', '=COUNTIF(الحركات!C2:C{},"JournalVoucher")'.format(n + 1))
    put(17, 'عدد الشيكات المرتجعة', '=COUNTIF(الحركات!C2:C{},"ReturnedCheque")'.format(n + 1))
    put(18, 'قيمة الشيكات المرتجعة (مدين)',
        '=SUMIF(الحركات!C2:C{0},"ReturnedCheque",الحركات!H2:H{0})'.format(n + 1),
        MONEY_FMT, bold=True)
    if receivables_df is not None:
        m = len(receivables_df)
        put(20, 'عدد بوالص ملف الذمم', m)
        put(21, 'صافي أرصدة الذمم المفتوحة',
            '=SUM(الذمم!M2:M{})'.format(m + 1), MONEY_FMT, bold=True)
        put(22, 'شيكات مرتجعة بملف الذمم',
            '=SUM(الذمم!L2:L{})'.format(m + 1), MONEY_FMT, bold=True)
    ws.cell(row=24, column=1,
            value='المصدر: كشف حساب ESKA بتاريخ ' + info.get('to', '') +
            ' + تقرير ذمم Crystal').font = Font(italic=True, size=9, name='Arial')

    wb.save(out_path)
    return out_path


def copy_to_icloud(source_path, icloud_folder='Ahleia Statements'):
    """Copy a generated file to iCloud Drive.

    :param source_path: Path to the local file to copy (str or Path).
    :param icloud_folder: Sub-folder name inside iCloud Drive (default: 'Ahleia Statements').
    :return: Destination Path on success, None if iCloud Drive is not available or copy fails.
    """
    icloud_drive = Path(os.environ.get(
        'ICLOUD_DRIVE_PATH',
        os.path.expanduser('~/Library/Mobile Documents/com~apple~CloudDocs'),
    ))
    if not icloud_drive.exists():
        print('WARNING: iCloud Drive not found at {}. Skipping iCloud export.'.format(icloud_drive))
        return None

    destination = icloud_drive / icloud_folder / Path(source_path).name
    try:
        destination.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy(source_path, destination)
    except OSError as exc:
        print('ERROR: Could not copy to iCloud Drive: {}'.format(exc))
        return None

    print('Copied to iCloud Drive: {}'.format(destination))
    return destination


def main():
    p = argparse.ArgumentParser(
        description='Generate an Excel workbook from an AHLEIA/ESKA PDF statement.'
    )
    p.add_argument('pdf', help='Path to the statement PDF file.')
    p.add_argument('--receivables', default=None,
                   help='Path to a Crystal receivables .xls/.xlsx file.')
    p.add_argument('-o', '--output', default=None,
                   help='Output .xlsx path (default: ahleia_<account>.xlsx).')
    p.add_argument('--icloud', action='store_true', default=False,
                   help='Copy the generated workbook to iCloud Drive (macOS only).')
    p.add_argument('--icloud-folder', default='Ahleia Statements',
                   help='Sub-folder inside iCloud Drive to copy the file into '
                        '(default: "Ahleia Statements"). Only used when --icloud is set.')
    a = p.parse_args()

    info, rows = parse_statement(a.pdf)
    rows = enrich(rows)
    if not rows:
        print('ERROR: No transaction rows parsed from the PDF. Aborting.')
        return
    print('Parsed {} rows | account {} | final balance {:,.2f}'.format(
        len(rows), info.get('account'), rows[-1]['balance']))

    rec = None
    if a.receivables:
        rec = pd.read_excel(a.receivables, header=1)
        rec = rec.dropna(subset=['DOCUMENT_NO'])
        print('Receivables: {} policies'.format(len(rec)))

    out = a.output or 'ahleia_{}.xlsx'.format(
        info.get('account', 'statement').replace('-', ''))
    build_workbook(info, rows, rec, out)
    print('Saved: {}'.format(out))

    if a.icloud:
        copy_to_icloud(out, icloud_folder=a.icloud_folder)


if __name__ == '__main__':
    main()
