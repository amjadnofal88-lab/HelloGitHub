#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
renewals_report.py — كشف التجديدات
يقرأ تقرير الذمم (Crystal .xls/.xlsx) ويولّد كشف تجديدات إكسل عربي RTL:
  • متأخرة   — انتهت ولم تُجدَّد
  • عاجلة    — تنتهي خلال 15 يوم
  • قادمة    — تنتهي خلال 16-60 يوم
  • الكل     — كل البوالص مع الحالة والأيام المتبقية
  • واتساب   — ملف جاهز للبوت (اسم / جوال / نص الرسالة)

Usage:
    python renewals_report.py receivables.xls
    python renewals_report.py receivables.xls --contacts contacts.xlsx --days 60 -o renewals.xlsx
    python renewals_report.py receivables.xls --cc 972 --cloud
    python renewals_report.py receivables.xls --production
    python renewals_report.py receivables.xls --cc 970593666668 --production --cloud
"""
import os
import re
import shutil
import argparse
import datetime as dt
import subprocess
import tempfile
from pathlib import Path

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ── إعدادات قابلة للتعديل ───────────────────────────────────────────────
URGENT_DAYS = 15          # "عاجل" = ينتهي خلال هذا العدد من الأيام
UPCOMING_DAYS = 60        # نطاق "قادم"
BROKER_PHONE = '0599432658'
BROKER_NAME = 'شركة البديل للخدمات العامة واللوجستية'
ICLOUD_FOLDER = 'AHLEIA Renewals'
COUNTRY_CODE = '972'   # 972 = واتساب الضفة | 970 = الترقيم الفلسطيني

HDR_FILL = PatternFill(fill_type='solid', fgColor='1F4E78')
HDR_FONT = Font(bold=True, color='FFFFFF', name='Arial', size=11)
BODY_FONT = Font(name='Arial', size=10)
MONEY = '#,##0.00;[Red]-#,##0.00;-'

FILLS = {
    'متأخرة': PatternFill(fill_type='solid', fgColor='FFC7CE'),   # أحمر فاتح
    'عاجلة': PatternFill(fill_type='solid', fgColor='FFEB9C'),    # أصفر
    'قادمة': PatternFill(fill_type='solid', fgColor='C6EFCE'),    # أخضر
    'سارية': PatternFill(fill_type='solid', fgColor='FFFFFF'),
}


def load_receivables(path):
    """يقرأ .xls (Crystal) عبر LibreOffice أو .xlsx مباشرة."""
    path = Path(path)
    if path.suffix.lower() == '.xls':
        tmp = Path(tempfile.mkdtemp())
        try:
            subprocess.run(['soffice', '--headless', '--convert-to', 'xlsx',
                            '--outdir', str(tmp), str(path)],
                           check=True, capture_output=True)
        except (subprocess.CalledProcessError, FileNotFoundError) as exc:
            raise SystemExit(
                'ERROR: LibreOffice is required to convert .xls files. '
                'Please install it (e.g. apt install libreoffice) or convert '
                'the file to .xlsx manually.\nDetails: {}'.format(exc)
            ) from exc
        path = tmp / (path.stem + '.xlsx')
    df = pd.read_excel(path, header=1)
    return df.dropna(subset=['DOCUMENT_NO'])


def normalize_phone(v, cc='972'):
    """يحوّل الرقم لصيغة دولية.
    يُرجع (رقم مُنسّق 00cc..., معرّف واتساب cc..., سبب الرفض إن وُجد).
    """
    if pd.isna(v) or not str(v).strip():
        return '', '', 'فارغ'
    d = re.sub(r'\D', '', str(v))
    if not d:
        return '', '', 'لا يحتوي أرقام'
    if d.startswith('00'):
        d = d[2:]
    # جرّد أي مقدمة دولية موجودة (970 أو 972) وأعد بناءها بالمطلوب
    for pre in ('970', '972'):
        if d.startswith(pre) and len(d) >= 11:
            d = d[len(pre):]
            break
    d = d.lstrip('0')
    if len(d) < 9:
        return '', '', f'ناقص {9 - len(d)} خانة'
    if len(d) > 9:
        return '', '', f'زائد {len(d) - 9} خانة'
    if not d.startswith('5'):
        return '', '', 'ليس رقم جوال (لا يبدأ بـ 5)'
    national = cc + d
    return '00' + national, national, ''


def base_policy(doc_no):
    """رقم البوليصة الأساسي بدون لواحق التجديد/الملاحق."""
    return re.split(r'-R-|-E-', str(doc_no))[0]


def build(df, contacts=None, days=UPCOMING_DAYS, today=None, cc=COUNTRY_CODE):
    today = today or dt.date.today()

    df = df.copy()
    df['EXPIRY'] = pd.to_datetime(df['EXPIRY_DATE'], errors='coerce')
    df['ISSUE'] = pd.to_datetime(df['ISSUE_DATE'], errors='coerce')
    df = df.dropna(subset=['EXPIRY'])
    df['BASE'] = df['DOCUMENT_NO'].map(base_policy)

    # نُبقي أحدث وثيقة لكل بوليصة (التجديد يلغي سابقه)
    df = df.sort_values('EXPIRY').drop_duplicates('BASE', keep='last')

    df['DAYS_LEFT'] = (df['EXPIRY'].dt.date - today).map(lambda x: x.days)

    def status(d):
        if d < 0:
            return 'متأخرة'
        if d <= URGENT_DAYS:
            return 'عاجلة'
        if d <= days:
            return 'قادمة'
        return 'سارية'

    df['STATUS'] = df['DAYS_LEFT'].map(status)

    rejects = []
    # دمج جهات الاتصال (اختياري)
    df['PHONE'] = ''
    df['WA_ID'] = ''
    if contacts is not None:
        cmap = {}
        name_col = next((c for c in contacts.columns
                         if str(c).upper() in ('NAME', 'BENEFICIARY', 'الاسم')), contacts.columns[0])
        ph_col = next((c for c in contacts.columns
                       if str(c).upper() in ('PHONE', 'MOBILE', 'الجوال', 'الهاتف')), contacts.columns[1])
        for _, r in contacts.iterrows():
            disp, wa, why = normalize_phone(r[ph_col], cc)
            if wa:
                cmap[str(r[name_col]).strip()] = (disp, wa)
            else:
                rejects.append([str(r[name_col]).strip(), str(r[ph_col]), why])
        pairs = df['BENEFICIARY'].map(lambda n: cmap.get(str(n).strip(), ('', '')))
        df['PHONE'] = [p[0] for p in pairs]
        df['WA_ID'] = [p[1] for p in pairs]

    # بوالص متابعة بلا رقم مطابق في قاعدة جهات الاتصال
    for _, r in df[(df['STATUS'] != 'سارية') & (df['WA_ID'] == '')].iterrows():
        rejects.append([r['BENEFICIARY'], '', 'لا يوجد رقم في قاعدة جهات الاتصال'])

    return df.sort_values('DAYS_LEFT'), rejects


def wa_message(row):
    d = int(row['DAYS_LEFT'])
    when = (f"انتهت بتاريخ {row['EXPIRY']:%d-%m-%Y} (متأخرة {abs(d)} يوم)"
            if d < 0 else f"تنتهي بتاريخ {row['EXPIRY']:%d-%m-%Y} (بعد {d} يوم)")
    veh = row.get('VEHICLE_TYPE') or ''
    plate = row.get('PLATE_NUMBER') or ''
    veh_txt = f' للمركبة {veh} لوحة {plate}'.rstrip() if str(plate) != 'nan' and plate else ''
    return (f"مرحباً {row['BENEFICIARY']}،\n"
            f"وثيقة التأمين رقم {row['DOCUMENT_NO']}{veh_txt} {when}.\n"
            f"يسعدنا تجديدها لك — تواصل معنا على {BROKER_PHONE}.\n"
            f"{BROKER_NAME}")


def sheet(wb, title, rows, headers, widths, money_cols=(), color_by_status=True):
    ws = wb.create_sheet(title)
    ws.sheet_view.rightToLeft = True
    ws.append(headers)
    for c in ws[1]:
        c.font, c.fill = HDR_FONT, HDR_FILL
        c.alignment = Alignment(horizontal='center')
    for r in rows:
        ws.append(r)
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for col in money_cols:
        for c in ws[get_column_letter(col)][1:]:
            c.number_format = MONEY
    if color_by_status and 'الحالة' in headers:
        si = headers.index('الحالة') + 1
        for row in ws.iter_rows(min_row=2):
            st = row[si - 1].value
            if st in FILLS:
                for c in row:
                    c.fill = FILLS[st]
                    c.font = BODY_FONT
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions
    return ws


COLS = ['BENEFICIARY', 'DOCUMENT_NO', 'CLASS', 'VEHICLE_TYPE', 'VEHICLE_MODEL',
        'PLATE_NUMBER', 'EXPIRY', 'DAYS_LEFT', 'STATUS', 'NET_PREMIUM_LC',
        'NET_AR_BALANCE', 'PHONE', 'WA_ID']
HEADERS = ['المؤمَّن له', 'رقم البوليصة', 'الفرع', 'نوع المركبة', 'الموديل',
           'رقم اللوحة', 'تاريخ الانتهاء', 'الأيام المتبقية', 'الحالة',
           'القسط', 'الرصيد المستحق', 'الجوال', 'معرّف واتساب']
WIDTHS = [30, 24, 12, 14, 14, 12, 14, 13, 11, 12, 14, 15, 15]


def rows_of(d):
    out = []
    for _, r in d.iterrows():
        out.append([
            r['BENEFICIARY'], r['DOCUMENT_NO'], r['CLASS'],
            r.get('VEHICLE_TYPE') if pd.notna(r.get('VEHICLE_TYPE')) else '',
            r.get('VEHICLE_MODEL') if pd.notna(r.get('VEHICLE_MODEL')) else '',
            str(r.get('PLATE_NUMBER')) if pd.notna(r.get('PLATE_NUMBER')) else '',
            r['EXPIRY'].strftime('%d-%m-%Y'), int(r['DAYS_LEFT']), r['STATUS'],
            float(r['NET_PREMIUM_LC']), float(r['NET_AR_BALANCE']),
            r['PHONE'], r['WA_ID'],
        ])
    return out


def add_production_sheets(wb, raw_df):
    """يضيف ثلاثة أوراق لتقرير الإنتاج إلى المصنف:
      • الإنتاج        — كل البوالص مرتبة بتاريخ الإصدار تنازلياً
      • الإنتاج حسب الفرع — ملخص القسط وعدد البوالص لكل فرع
      • الإنتاج الشهري  — ملخص شهري
    يعمل على raw_df (قبل إزالة التكرارات) لإظهار كل الإصدارات.
    """
    df = raw_df.copy()
    df['ISSUE'] = pd.to_datetime(df.get('ISSUE_DATE'), errors='coerce')
    df['EXPIRY'] = pd.to_datetime(df.get('EXPIRY_DATE'), errors='coerce')
    df = df.dropna(subset=['ISSUE'])
    df = df.sort_values('ISSUE', ascending=False)

    # ── شيت الإنتاج الكامل ─────────────────────────────────────────────
    p_hdrs = ['تاريخ الإصدار', 'تاريخ الانتهاء', 'المؤمَّن له',
              'رقم البوليصة', 'الفرع', 'نوع المركبة', 'الموديل',
              'رقم اللوحة', 'القسط']
    p_widths = [14, 14, 30, 24, 12, 14, 14, 12, 13]
    p_rows = []
    for _, r in df.iterrows():
        p_rows.append([
            r['ISSUE'].strftime('%d-%m-%Y'),
            r['EXPIRY'].strftime('%d-%m-%Y') if pd.notna(r.get('EXPIRY')) else '',
            r['BENEFICIARY'],
            r['DOCUMENT_NO'],
            r.get('CLASS', ''),
            r.get('VEHICLE_TYPE', '') if pd.notna(r.get('VEHICLE_TYPE')) else '',
            r.get('VEHICLE_MODEL', '') if pd.notna(r.get('VEHICLE_MODEL')) else '',
            str(r.get('PLATE_NUMBER', '')) if pd.notna(r.get('PLATE_NUMBER')) else '',
            float(r['NET_PREMIUM_LC']),
        ])
    sheet(wb, 'الإنتاج', p_rows, p_hdrs, p_widths, money_cols=(9,), color_by_status=False)

    # ── شيت الإنتاج حسب الفرع ─────────────────────────────────────────
    by_class = (df.groupby('CLASS', dropna=False)
                  .agg(عدد_البوالص=('DOCUMENT_NO', 'count'),
                       إجمالي_الأقساط=('NET_PREMIUM_LC', 'sum'))
                  .reset_index()
                  .sort_values('إجمالي_الأقساط', ascending=False))
    cls_rows = [[r['CLASS'], int(r['عدد_البوالص']), float(r['إجمالي_الأقساط'])]
                for _, r in by_class.iterrows()]
    # صف الإجمالي
    cls_rows.append(['الإجمالي', len(df), float(df['NET_PREMIUM_LC'].sum())])
    ws_cls = sheet(wb, 'الإنتاج حسب الفرع', cls_rows,
                   ['الفرع', 'عدد البوالص', 'إجمالي الأقساط'],
                   [20, 16, 18], money_cols=(3,), color_by_status=False)
    # تمييز صف الإجمالي
    total_row = ws_cls.max_row
    for c in ws_cls[total_row]:
        c.font = Font(bold=True, name='Arial', size=10)
        c.fill = PatternFill(fill_type='solid', fgColor='D9E1F2')

    # ── شيت الإنتاج الشهري ────────────────────────────────────────────
    df['شهر'] = df['ISSUE'].dt.to_period('M')
    by_month = (df.groupby('شهر', dropna=False)
                  .agg(عدد_البوالص=('DOCUMENT_NO', 'count'),
                       إجمالي_الأقساط=('NET_PREMIUM_LC', 'sum'))
                  .reset_index()
                  .sort_values('شهر', ascending=False))
    mo_rows = [[str(r['شهر']), int(r['عدد_البوالص']), float(r['إجمالي_الأقساط'])]
               for _, r in by_month.iterrows()]
    mo_rows.append(['الإجمالي', len(df), float(df['NET_PREMIUM_LC'].sum())])
    ws_mo = sheet(wb, 'الإنتاج الشهري', mo_rows,
                  ['الشهر', 'عدد البوالص', 'إجمالي الأقساط'],
                  [14, 16, 18], money_cols=(3,), color_by_status=False)
    total_row = ws_mo.max_row
    for c in ws_mo[total_row]:
        c.font = Font(bold=True, name='Arial', size=10)
        c.fill = PatternFill(fill_type='solid', fgColor='D9E1F2')


def build_workbook(df, out_path, days=UPCOMING_DAYS, rejects=None, raw_df=None):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    late = df[df['STATUS'] == 'متأخرة']
    urgent = df[df['STATUS'] == 'عاجلة']
    soon = df[df['STATUS'] == 'قادمة']
    action = df[df['STATUS'] != 'سارية']

    # ملخص
    ws = wb.create_sheet('الملخص')
    ws.sheet_view.rightToLeft = True
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18
    ws['A1'] = 'كشف التجديدات'
    ws['A1'].font = Font(bold=True, size=14, name='Arial', color='1F4E78')
    ws['A2'] = f'تاريخ التقرير: {dt.date.today():%d-%m-%Y}'
    ws['A2'].font = Font(italic=True, size=9, name='Arial')

    ws.append([])
    hdr = ['الحالة', 'عدد البوالص', 'إجمالي الأقساط']
    ws.append(hdr)
    for c in ws[4]:
        c.font, c.fill = HDR_FONT, HDR_FILL
    for name, d in [('متأخرة', late), ('عاجلة (\u226415 يوم)', urgent),
                    (f'قادمة (\u2264{days} يوم)', soon),
                    ('إجمالي للمتابعة', action), ('سارية', df[df['STATUS'] == 'سارية'])]:
        ws.append([name, len(d), float(d['NET_PREMIUM_LC'].sum())])
    for row in ws.iter_rows(min_row=5, min_col=3, max_col=3):
        for c in row:
            c.number_format = MONEY
            c.font = BODY_FONT
    ws.cell(row=9, column=1).font = Font(bold=True, name='Arial')
    ws.cell(row=9, column=2).font = Font(bold=True, name='Arial')
    ws.cell(row=9, column=3).font = Font(bold=True, name='Arial')
    ws.append([])
    ws.append(['أرصدة مستحقة على بوالص المتابعة', '',
               float(action['NET_AR_BALANCE'].sum())])
    ws.cell(row=ws.max_row, column=3).number_format = MONEY
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True, name='Arial')

    sheet(wb, 'متأخرة', rows_of(late), HEADERS, WIDTHS, money_cols=(10, 11))
    sheet(wb, 'عاجلة', rows_of(urgent), HEADERS, WIDTHS, money_cols=(10, 11))
    sheet(wb, 'قادمة', rows_of(soon), HEADERS, WIDTHS, money_cols=(10, 11))
    sheet(wb, 'الكل', rows_of(df), HEADERS, WIDTHS, money_cols=(10, 11))

    # شيت واتساب — جاهز للبوت
    wa_rows = [[r['BENEFICIARY'], r['PHONE'], r['WA_ID'], r['DOCUMENT_NO'],
                r['EXPIRY'].strftime('%d-%m-%Y'), int(r['DAYS_LEFT']),
                r['STATUS'], wa_message(r)] for _, r in action.iterrows()]
    ws_wa = sheet(wb, 'واتساب', wa_rows,
                  ['الاسم', 'الجوال', 'معرّف واتساب', 'رقم البوليصة',
                   'تاريخ الانتهاء', 'الأيام المتبقية', 'الحالة', 'نص الرسالة'],
                  [28, 16, 16, 24, 14, 13, 11, 90], color_by_status=False)
    for col in ('B', 'C'):                       # نص، حتى لا يحذف إكسل الأصفار
        for c in ws_wa[col][1:]:
            c.number_format = '@'

    # شيت الأرقام غير الصالحة
    if rejects:
        ws_bad = sheet(wb, 'أرقام غير صالحة', rejects,
                       ['الاسم', 'الرقم كما ورد', 'سبب الرفض'],
                       [32, 22, 30], color_by_status=False)
        bad_fill = PatternFill(fill_type='solid', fgColor='FFC7CE')
        for row in ws_bad.iter_rows(min_row=2):
            for c in row:
                c.fill = bad_fill
                c.font = BODY_FONT
            row[1].number_format = '@'

    # أوراق تقرير الإنتاج (اختياري)
    if raw_df is not None:
        add_production_sheets(wb, raw_df)

    wb.save(out_path)
    return out_path


def copy_to_icloud(src, folder=ICLOUD_FOLDER):
    root = Path(os.environ.get(
        'ICLOUD_DRIVE_PATH',
        str(Path.home() / 'Library' / 'Mobile Documents' / 'com~apple~CloudDocs')))
    if not root.exists():
        print(f'WARNING: iCloud Drive not found at {root}. Skipping.')
        return None
    dest = root / folder / Path(src).name
    try:
        dest.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy(src, dest)
    except OSError as e:
        print(f'ERROR: iCloud copy failed: {e}')
        return None
    print(f'Copied to iCloud: {dest}')
    return dest


def main():
    p = argparse.ArgumentParser(description='كشف تجديدات البوالص')
    p.add_argument('receivables', help='ملف الذمم (.xls أو .xlsx)')
    p.add_argument('--contacts', help='ملف جهات الاتصال (اسم + جوال) لدمج الأرقام')
    p.add_argument('--days', type=int, default=UPCOMING_DAYS,
                   help=f'نطاق التجديدات القادمة بالأيام (افتراضي {UPCOMING_DAYS})')
    p.add_argument('-o', '--output', default=None)
    p.add_argument('--cc', default=COUNTRY_CODE, choices=['972', '970'],
                   help='رمز الدولة للأرقام (افتراضي 972 — صيغة واتساب)')
    p.add_argument('--icloud', '--cloud', dest='icloud', action='store_true',
                   help='نسخ الملف إلى iCloud Drive')
    p.add_argument('--icloud-folder', default=ICLOUD_FOLDER)
    p.add_argument('--production', action='store_true',
                   help='إضافة أوراق تقرير الإنتاج (الإنتاج / حسب الفرع / الشهري)')
    a = p.parse_args()

    raw_df = load_receivables(a.receivables)
    contacts = pd.read_excel(a.contacts) if a.contacts else None
    df, rejects = build(raw_df, contacts=contacts, days=a.days, cc=a.cc)

    out = a.output or f'renewals_{dt.date.today():%Y%m%d}.xlsx'
    build_workbook(df, out, days=a.days, rejects=rejects,
                   raw_df=raw_df if a.production else None)

    c = df['STATUS'].value_counts()
    print(f"Saved: {out}")
    print(f"  متأخرة: {c.get('متأخرة', 0)} | عاجلة: {c.get('عاجلة', 0)} | "
          f"قادمة: {c.get('قادمة', 0)} | سارية: {c.get('سارية', 0)}")
    if rejects:
        print(f"  أرقام غير صالحة / مفقودة: {len(rejects)}")

    if a.icloud:
        copy_to_icloud(out, a.icloud_folder)


if __name__ == '__main__':
    main()
