#!/usr/bin/env python
# -*- coding:utf-8 -*-
"""
Fetch customer data files from a GitHub repository and export them to Excel.
"""
import argparse
import base64
import csv
import io
import json
import os
import sqlite3
import tempfile
from datetime import datetime

import requests
from openpyxl import Workbook, load_workbook

GITHUB_API = 'https://api.github.com'
DEFAULT_OWNER = 'amjadnofal88-lab'
DEFAULT_REPO = 'HelloGitHub'
DEFAULT_BRANCH = 'main'
SUPPORTED_EXTENSIONS = {'.xlsx', '.xls', '.csv', '.json', '.db', '.sqlite'}


def build_headers(token=None):
    """Build GitHub API headers."""
    headers = {'Accept': 'application/vnd.github.v3+json'}
    if token:
        headers['Authorization'] = 'token {}'.format(token)
    return headers


def list_repo_files(owner, repo, path='', branch=DEFAULT_BRANCH, headers=None):
    """List files in a GitHub repository path."""
    url = '{}/repos/{}/{}/contents/{}'.format(GITHUB_API, owner, repo, path)
    response = requests.get(
        url,
        headers=headers or {},
        params={'ref': branch},
        timeout=10,
    )
    if response.status_code != 200:
        message = response.json().get('message', 'Unknown error')
        print('❌ تعذر قراءة {}: {} — {}'.format(path or '/', response.status_code, message))
        return []
    return response.json()


def download_file(owner, repo, file_path, branch=DEFAULT_BRANCH, headers=None):
    """Download a file from GitHub and return raw bytes."""
    url = '{}/repos/{}/{}/contents/{}'.format(GITHUB_API, owner, repo, file_path)
    response = requests.get(
        url,
        headers=headers or {},
        params={'ref': branch},
        timeout=10,
    )
    if response.status_code != 200:
        print('❌ فشل تحميل {}: {}'.format(file_path, response.status_code))
        return None

    payload = response.json()
    if payload.get('encoding') != 'base64':
        print('⚠️ تنسيق غير مدعوم للملف {}'.format(file_path))
        return None
    return base64.b64decode(payload['content'])


def find_data_files(owner, repo, branch=DEFAULT_BRANCH, path='', headers=None, results=None):
    """Recursively find supported data files in the repository."""
    if results is None:
        results = []

    items = list_repo_files(owner, repo, path=path, branch=branch, headers=headers)
    for item in items:
        item_type = item.get('type')
        if item_type == 'file':
            ext = os.path.splitext(item['name'])[1].lower()
            if ext in SUPPORTED_EXTENSIONS:
                results.append(item)
                print('   ✅ وجدت: {} ({} bytes)'.format(item['path'], item.get('size', 0)))
        elif item_type == 'dir':
            find_data_files(owner, repo, branch=branch, path=item['path'], headers=headers, results=results)
    return results


def _normalize_key(value, index):
    text = '' if value is None else str(value).strip()
    return text or 'column_{}'.format(index)


def _rows_to_dicts(rows):
    if not rows:
        return []
    headers = [_normalize_key(value, index + 1) for index, value in enumerate(rows[0])]
    records = []
    for values in rows[1:]:
        if not any(value is not None and value != '' for value in values):
            continue
        record = {}
        for index, header in enumerate(headers):
            record[header] = values[index] if index < len(values) else None
        records.append(record)
    return records


def extract_csv_records(content):
    """Extract records from CSV bytes."""
    text = None
    for encoding in ('utf-8', 'utf-8-sig', 'cp1256'):
        try:
            text = content.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        raise ValueError('تعذر قراءة CSV بترميز معروف')
    reader = csv.DictReader(io.StringIO(text))
    return [dict(row) for row in reader]


def extract_json_records(content):
    """Extract records from JSON bytes."""
    data = json.loads(content.decode('utf-8'))
    if isinstance(data, list):
        return [item if isinstance(item, dict) else {'value': item} for item in data]
    if isinstance(data, dict):
        preferred_lists = []
        fallback_lists = []
        for key, value in data.items():
            if isinstance(value, list) and value:
                if 'customer' in key.lower():
                    preferred_lists.append(value)
                else:
                    fallback_lists.append(value)
        if preferred_lists:
            return [item if isinstance(item, dict) else {'value': item} for item in preferred_lists[0]]
        if fallback_lists:
            return [item if isinstance(item, dict) else {'value': item} for item in fallback_lists[0]]
        return [data]
    raise ValueError('تنسيق JSON غير مدعوم')


def extract_xlsx_records(content):
    """Extract records from an XLSX workbook."""
    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    records = []
    for worksheet in workbook.worksheets:
        rows = list(worksheet.iter_rows(values_only=True))
        for record in _rows_to_dicts(rows):
            record['_source_sheet'] = worksheet.title
            records.append(record)
    return records


def _candidate_tables(table_names):
    if 'customers' in table_names:
        return ['customers']
    matching = [name for name in table_names if 'customer' in name.lower()]
    if matching:
        return matching
    if len(table_names) == 1:
        return table_names
    return []


def extract_sqlite_records(content):
    """Extract records from a SQLite database dump."""
    temp_path = None
    try:
        with tempfile.NamedTemporaryFile(suffix='.db', delete=False) as temp_file:
            temp_file.write(content)
            temp_path = temp_file.name

        connection = sqlite3.connect(temp_path)
        connection.row_factory = sqlite3.Row
        try:
            tables = connection.execute(
                "SELECT name FROM sqlite_master WHERE type = 'table' AND name NOT LIKE 'sqlite_%'"
            ).fetchall()
            table_names = [row['name'] for row in tables]
            selected_tables = _candidate_tables(table_names)
            records = []
            for table_name in selected_tables:
                rows = connection.execute('SELECT * FROM "{}"'.format(table_name)).fetchall()
                for row in rows:
                    record = dict(row)
                    record['_source_table'] = table_name
                    records.append(record)
            return records
        finally:
            connection.close()
    finally:
        if temp_path and os.path.exists(temp_path):
            os.unlink(temp_path)


def process_file(content, file_info):
    """Dispatch file processing by extension."""
    ext = os.path.splitext(file_info['name'])[1].lower()
    if ext == '.csv':
        return extract_csv_records(content), 'csv'
    if ext == '.json':
        return extract_json_records(content), 'json'
    if ext == '.xlsx':
        return extract_xlsx_records(content), 'xlsx'
    if ext == '.xls':
        raise ValueError('صيغة .xls غير مدعومة، حوّل الملف إلى .xlsx أولاً')
    if ext in ('.db', '.sqlite'):
        return extract_sqlite_records(content), 'sqlite'
    raise ValueError('صيغة غير مدعومة: {}'.format(ext))


def enrich_records(records, file_info, source_type):
    """Attach source metadata to extracted records."""
    enriched = []
    for record in records:
        item = dict(record)
        item['_source_file'] = file_info['path']
        item['_source_type'] = source_type
        enriched.append(item)
    return enriched


def deduplicate_records(records):
    """Remove duplicate records while preserving order."""
    unique = []
    seen = set()
    for record in records:
        key = json.dumps(record, ensure_ascii=False, sort_keys=True, default=str)
        if key in seen:
            continue
        seen.add(key)
        unique.append(record)
    return unique


def export_to_excel(records, output_path):
    """Export records to an Excel file."""
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'customers'

    if not records:
        worksheet.append(['message'])
        worksheet.append(['No customer records found'])
    else:
        headers = []
        for record in records:
            for key in record.keys():
                if key not in headers:
                    headers.append(key)
        worksheet.append(headers)
        for record in records:
            worksheet.append([record.get(header) for header in headers])

    workbook.save(output_path)
    return output_path


def fetch_customer_records(owner, repo, branch=DEFAULT_BRANCH, token=None):
    """Fetch and process supported data files from a repository."""
    headers = build_headers(token)
    data_files = find_data_files(owner, repo, branch=branch, headers=headers)
    all_records = []

    for file_info in data_files:
        print('\n📄 {}'.format(file_info['path']))
        content = download_file(owner, repo, file_info['path'], branch=branch, headers=headers)
        if not content:
            continue
        try:
            records, source_type = process_file(content, file_info)
        except Exception as exc:
            print('   ❌ خطأ في المعالجة: {}'.format(exc))
            continue

        if not records:
            print('   ⚠️ لا توجد سجلات قابلة للاستخراج')
            continue

        enriched = enrich_records(records, file_info, source_type)
        print('   📊 {} سجل'.format(len(enriched)))
        all_records.extend(enriched)

    return all_records


def parse_args():
    parser = argparse.ArgumentParser(
        description='Fetch customer data files from a GitHub repository and export them to Excel.'
    )
    parser.add_argument('--owner', default=DEFAULT_OWNER, help='GitHub owner/user name.')
    parser.add_argument('--repo', default=DEFAULT_REPO, help='GitHub repository name.')
    parser.add_argument('--branch', default=DEFAULT_BRANCH, help='Git branch or ref.')
    parser.add_argument(
        '--token',
        default=os.environ.get('GITHUB_TOKEN', ''),
        help='GitHub token. Defaults to the GITHUB_TOKEN environment variable.',
    )
    parser.add_argument(
        '--output',
        default='customers_{}.xlsx'.format(datetime.now().strftime('%Y%m%d_%H%M')),
        help='Output Excel file path.',
    )
    return parser.parse_args()


def main():
    args = parse_args()

    print('=' * 55)
    print('📂 GitHub Customer Data Fetcher')
    print('   Repo: {}/{}'.format(args.owner, args.repo))
    print('=' * 55)

    headers = build_headers(args.token)
    repo_url = '{}/repos/{}/{}'.format(GITHUB_API, args.owner, args.repo)
    response = requests.get(repo_url, headers=headers, timeout=10)
    if response.status_code != 200:
        print('\n❌ فشل الاتصال بالمستودع: {}'.format(response.json().get('message', response.status_code)))
        return 1

    print('✅ تم الاتصال بـ {}'.format(response.json().get('full_name', '{}/{}'.format(args.owner, args.repo))))
    print('\n🔍 جاري البحث عن ملفات البيانات...')
    records = fetch_customer_records(args.owner, args.repo, branch=args.branch, token=args.token)

    if not records:
        print('\n⚠️ لم يتم العثور على ملفات بيانات عملاء قابلة للمعالجة.')
        return 0

    unique_records = deduplicate_records(records)
    print('\n🧹 إزالة التكرارات: {} → {} سجل'.format(len(records), len(unique_records)))
    export_to_excel(unique_records, args.output)
    print('\n✅ تم التصدير: {}'.format(args.output))
    print('   📊 {} سجل'.format(len(unique_records)))
    return 0


if __name__ == '__main__':
    raise SystemExit(main())
