#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""eska_contacts.py — fetch policy and contact details from the Eska portal.

This script is intentionally written around the actual search form labels that the
portal exposes and keeps the credentials in environment variables only.

Examples:
    export ESKA_USER=...
    export ESKA_PASS=...

    python eska_contacts.py --from 01-01-2026 --to 31-12-2026
    python eska_contacts.py --from 01-01-2026 --to 31-12-2026 --renewals-only
    python eska_contacts.py --policies-file policies.txt
    python eska_contacts.py --from 01-01-2026 --to 31-12-2026 --no-headless
"""

import argparse
import datetime as dt
import json
import logging
import os
import random
import re
import sys
import time
from pathlib import Path

try:
    from playwright.sync_api import sync_playwright, TimeoutError as PWTimeout
except Exception:  # pragma: no cover - Playwright is optional for import-time tests.
    sync_playwright = None

    class PWTimeout(Exception):
        """Fallback timeout for environments without Playwright."""


BASE_URL = os.environ.get("ESKA_BASE_URL", "http://82.213.38.146")
LOGIN_PATH = os.environ.get("ESKA_LOGIN_PATH", "/")
SEARCH_PATH = os.environ.get("ESKA_SEARCH_PATH", "/")

DATE_FMT = "%d-%m-%Y"
PAGE_SIZE = "100"
MIN_DELAY = 1.2
MAX_DELAY = 2.8
MAX_RETRIES = 3
NAV_TIMEOUT = 45_000
CHECKPOINT = Path("eska_checkpoint.json")

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-7s  %(message)s",
    datefmt="%H:%M:%S",
)
logger = logging.getLogger("eska")


def polite_sleep():
    """Pause briefly to reduce the chance that requests look automated."""
    time.sleep(random.uniform(MIN_DELAY, MAX_DELAY))


def parse_date(value):
    """Convert a date-like value into a date object or return None if parsing fails."""
    if value is None:
        return None
    value = str(value).strip()
    if not value:
        return None
    for fmt in (DATE_FMT, "%d/%m/%Y", "%Y-%m-%d"):
        try:
            return dt.datetime.strptime(value, fmt).date()
        except ValueError:
            continue
    return None


def load_checkpoint():
    if CHECKPOINT.exists():
        try:
            return json.loads(CHECKPOINT.read_text(encoding="utf-8"))
        except (json.JSONDecodeError, OSError):
            logger.warning("تعذرت قراءة ملف الاستئناف — سيتم تجاهله.")
    return {"done": [], "rows": []}


def save_checkpoint(state):
    CHECKPOINT.write_text(json.dumps(state, ensure_ascii=False, indent=1), encoding="utf-8")


def with_retries(fn, what):
    """Execute fn() with a limited retry loop for timeouts and transient errors."""
    for attempt in range(1, MAX_RETRIES + 1):
        try:
            return fn()
        except PWTimeout:
            logger.warning("مهلة عند %s (محاولة %d/%d)", what, attempt, MAX_RETRIES)
        except Exception as exc:  # pragma: no cover - broad safety net.
            logger.warning("خطأ عند %s: %s (محاولة %d/%d)", what, exc, attempt, MAX_RETRIES)
        time.sleep(2 ** attempt)
    logger.error("فشل نهائي عند %s", what)
    return None


def field_by_label(page, label):
    """Return the first input that follows the label text in the page markup."""
    if not label:
        return page.locator("input").first
    labels = [label]
    normalized = label.replace(" ", "")
    if normalized != label:
        labels.append(normalized)
    labels.extend([label.lower(), label.title(), label.upper()])

    for candidate in labels:
        for xpath in (
            f"//*[normalize-space(text())='{candidate}']/following::input[not(@type='hidden')][1]",
            f"//*[contains(normalize-space(.), '{candidate}')]/following::input[not(@type='hidden')][1]",
        ):
            loc = page.locator(f"xpath={xpath}").first
            if loc.count():
                return loc
    return page.locator("input").first


def login(page, user, password):
    logger.info("تسجيل الدخول ...")
    page.goto(BASE_URL + LOGIN_PATH, timeout=NAV_TIMEOUT)
    page.wait_for_load_state("domcontentloaded")

    for label in ("User Name", "Username", "اسم المستخدم", "UserName"):
        loc = field_by_label(page, label)
        if loc.count():
            loc.fill(user or "")
            break

    for label in ("Password", "كلمة المرور", "Pass"):
        loc = field_by_label(page, label)
        if loc.count():
            loc.fill(password or "")
            break

    page.get_by_role("button", name=re.compile("login|sign in|دخول", re.I)).first.click()
    page.wait_for_load_state("networkidle", timeout=NAV_TIMEOUT)
    logger.info("تم الدخول.")


def search(
    page,
    *,
    date_from=None,
    date_to=None,
    policy_no=None,
    vehicle_make=None,
    vehicle_type=None,
    plate_no=None,
    customer_id=None,
    sequence_no=None,
    posting_status=None,
):
    """Fill the search form and trigger the portal search."""
    if SEARCH_PATH not in page.url:
        page.goto(BASE_URL + SEARCH_PATH, timeout=NAV_TIMEOUT)
        page.wait_for_load_state("domcontentloaded")

    mapping = {
        "From Issue Date": date_from,
        "To Issue Date": date_to,
        "Policy No": policy_no,
        "Vehicle Make": vehicle_make,
        "Vihicle Make": vehicle_make,
        "Vehicle Type": vehicle_type,
        "Vihicle Type": vehicle_type,
        "Plate No": plate_no,
        "Custormer ID": customer_id,
        "Customer ID": customer_id,
        "Sequence No": sequence_no,
        "Posting Status": posting_status,
    }
    for label, value in mapping.items():
        if value in (None, ""):
            continue
        loc = field_by_label(page, label)
        if loc.count():
            loc.fill("")
            loc.type(str(value), delay=40)
        else:
            logger.warning("لم يعثر على الحقل: %s", label)

    page.get_by_role("button", name=re.compile(r"^\s*search\s*$", re.I)).first.click()
    page.wait_for_load_state("networkidle", timeout=NAV_TIMEOUT)
    polite_sleep()


def set_page_size(page):
    """Increase the number of items shown in the policy list when available."""
    show = page.locator("xpath=//*[contains(text(),'Show')]/following::select[1]").first
    if not show.count():
        return
    options = [o.strip() for o in show.locator("option").all_inner_texts()]
    target = PAGE_SIZE if PAGE_SIZE in options else (options[-1] if options else None)
    if target:
        show.select_option(target)
        page.wait_for_load_state("networkidle", timeout=NAV_TIMEOUT)
        logger.info("عدد الصفوف المعروضة: %s", target)


def read_policy_list(page):
    """Read the policy list table and return rows as dictionaries."""
    table = page.locator("xpath=//*[contains(text(),'Policy List')]/following::table[1]").first
    if not table.count():
        found = page.locator("table").first
        if found.count():
            table = found
        else:
            logger.warning("جدول Policy List غير موجود في هذه الصفحة.")
            return []

    rows = []
    headers = []
    for tr in table.locator("tr").all():
        cells = [normalize_text(c) for c in tr.locator("td, th").all_inner_texts()]
        if not cells or not any(cells):
            continue
        if not headers:
            headers = cells
            continue
        record = dict(zip(headers, cells))
        policy_no = cells[0] if len(cells) > 0 else ""
        effective_date = cells[1] if len(cells) > 1 else ""
        expiry_date = cells[2] if len(cells) > 2 else ""
        customer_ref = cells[3] if len(cells) > 3 else ""
        record["policy_no"] = policy_no
        record["effective_date"] = effective_date
        record["expiry_date"] = expiry_date
        record["customer_ref"] = customer_ref
        if record.get("policy_no"):
            rows.append(record)
    return rows


def read_contact(page, policy_no):
    """Open a policy detail view and read the customer contact fields."""
    link = page.get_by_text(policy_no, exact=True).first
    if not link.count():
        return {}
    link.click()
    page.wait_for_load_state("networkidle", timeout=NAV_TIMEOUT)
    polite_sleep()

    contact = {}
    for key, labels in {
        "name": ("Customer Name", "Insured Name", "اسم العميل"),
        "phone": ("Phone", "Tel", "هاتف"),
        "mobile": ("Mobile", "Cell", "جوال"),
        "email": ("Email", "E-Mail", "بريد", "Email Address"),
        "id_no": ("ID No", "Identity", "رقم الهوية"),
    }.items():
        for label in labels:
            loc = field_by_label(page, label)
            if loc.count():
                value = normalize_text(loc.input_value() or "")
                if value:
                    contact[key] = value
                    break
    page.go_back(timeout=NAV_TIMEOUT)
    page.wait_for_load_state("networkidle", timeout=NAV_TIMEOUT)
    return contact


def normalize_text(value):
    return " ".join(str(value).split()) if value is not None else ""


def export(rows, out_path):
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError as exc:  # pragma: no cover - handled by runtime requirement.
        raise RuntimeError("openpyxl is required to export Excel files") from exc

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "contacts"
    ws.sheet_view.rightToLeft = True

    headers = [
        "رقم الوثيقة",
        "تاريخ السريان",
        "تاريخ الانتهاء",
        "أيام متبقية",
        "اسم العميل",
        "هاتف",
        "جوال",
        "بريد إلكتروني",
        "رقم الهوية",
        "مرجع العميل",
    ]
    fill = PatternFill(fill_type="solid", start_color="1F4E78", end_color="1F4E78")
    for col, text in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=text)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center")

    today = dt.date.today()
    for r, row in enumerate(rows, start=2):
        expiry = parse_date(row.get("expiry_date"))
        remaining = (expiry - today).days if expiry else ""
        values = [
            row.get("policy_no", ""),
            row.get("effective_date", ""),
            row.get("expiry_date", ""),
            remaining,
            row.get("name", ""),
            row.get("phone", ""),
            row.get("mobile", ""),
            row.get("email", ""),
            row.get("id_no", ""),
            row.get("customer_ref", ""),
        ]
        for c, value in enumerate(values, start=1):
            ws.cell(row=r, column=c, value=value)

    for col, width in zip("ABCDEFGHIJ", [26, 14, 14, 11, 30, 16, 16, 30, 16, 20]):
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A2"
    wb.save(out_path)
    logger.info("حُفظ الملف: %s (%d صف)", out_path, len(rows))


def main():
    parser = argparse.ArgumentParser(description="سحب بيانات الوثائق من بوابة إسكا")
    parser.add_argument("--from", dest="date_from", help="From Issue Date (DD-MM-YYYY)")
    parser.add_argument("--to", dest="date_to", help="To Issue Date (DD-MM-YYYY)")
    parser.add_argument("--policies-file", help="ملف نصي فيه رقم وثيقة في كل سطر (وضع احتياطي)")
    parser.add_argument("--policy-no", help="رقم وثيقة واحد")
    parser.add_argument("--plate-no", help="Plate No")
    parser.add_argument("--customer-id", help="Customer ID")
    parser.add_argument("--vehicle-make", help="Vehicle Make")
    parser.add_argument("--vehicle-type", help="Vehicle Type")
    parser.add_argument("--sequence-no", help="Sequence No")
    parser.add_argument("--posting-status", help="Posting Status")
    parser.add_argument("--renewals-only", action="store_true", help="الوثائق المنتهية خلال المدة المحددة فقط")
    parser.add_argument("--window", type=int, default=90, help="نافذة التجديد بالأيام (افتراضي ٩٠)")
    parser.add_argument("--out", default="contacts.xlsx")
    parser.add_argument("--no-headless", action="store_true")
    parser.add_argument("--resume", action="store_true", help="استئناف من آخر نقطة توقف")
    args = parser.parse_args()

    user = os.environ.get("ESKA_USER")
    password = os.environ.get("ESKA_PASS")
    if not password:
        sys.exit("ضع كلمة المرور في متغير البيئة ESKA_PASS قبل التشغيل.")

    has_filter = bool(args.date_from or args.date_to or args.policies_file or args.policy_no or args.plate_no or args.customer_id or args.vehicle_make or args.vehicle_type or args.sequence_no or args.posting_status)
    if not has_filter:
        sys.exit("حدّد مدى تاريخ عبر --from/--to أو ملف وثائق عبر --policies-file أو حقل بحث فردي.")

    if BASE_URL.startswith("http://"):
        logger.warning("البوابة تعمل على HTTP بدون تشفير — بيانات الدخول تُرسل نصاً مكشوفاً. استخدم شبكة موثوقة.")

    if sync_playwright is None:
        sys.exit("Playwright غير مثبت. قم بتثبيت المتطلبات قبل تشغيل هذا السكربت.")

    state = load_checkpoint() if args.resume else {"done": [], "rows": []}
    done = set(state.get("done", []))
    rows = list(state.get("rows", []))

    with sync_playwright() as pw:
        browser = pw.chromium.launch(headless=not args.no_headless)
        page = browser.new_page()
        page.set_default_timeout(NAV_TIMEOUT)

        try:
            login(page, user or "", password)

            if args.policies_file:
                policies = [
                    line.strip()
                    for line in Path(args.policies_file).read_text(encoding="utf-8").splitlines()
                    if line.strip()
                ]
                logger.info("وضع البحث الفردي — %d وثيقة.", len(policies))
                for i, policy_no in enumerate(policies, start=1):
                    if policy_no in done:
                        continue
                    logger.info("[%d/%d] %s", i, len(policies), policy_no)
                    ok = with_retries(
                        lambda p=policy_no: search(page, policy_no=p),
                        f"البحث عن {policy_no}",
                    )
                    if ok is None:
                        continue
                    found = read_policy_list(page)
                    for record in found:
                        record.update(read_contact(page, record.get("policy_no", "")) or {})
                        rows.append(record)
                    done.add(policy_no)
                    save_checkpoint({"done": sorted(done), "rows": rows})
            else:
                logger.info("البحث بمدى التاريخ: %s ← %s", args.date_from, args.date_to)
                with_retries(
                    lambda: search(
                        page,
                        date_from=args.date_from,
                        date_to=args.date_to,
                        policy_no=args.policy_no,
                        vehicle_make=args.vehicle_make,
                        vehicle_type=args.vehicle_type,
                        plate_no=args.plate_no,
                        customer_id=args.customer_id,
                        sequence_no=args.sequence_no,
                        posting_status=args.posting_status,
                    ),
                    "البحث بمدى التاريخ",
                )
                set_page_size(page)
                found = read_policy_list(page)
                logger.info("عدد الوثائق في القائمة: %d", len(found))

                for i, record in enumerate(found, start=1):
                    policy_no = record.get("policy_no")
                    if not policy_no or policy_no in done:
                        continue
                    logger.info("[%d/%d] %s", i, len(found), policy_no)
                    contact = with_retries(
                        lambda p=policy_no: read_contact(page, p),
                        f"قراءة تفاصيل {policy_no}",
                    )
                    record.update(contact or {})
                    rows.append(record)
                    done.add(policy_no)
                    save_checkpoint({"done": sorted(done), "rows": rows})
        finally:
            browser.close()

    if args.renewals_only:
        today = dt.date.today()
        limit = today + dt.timedelta(days=args.window)
        before = len(rows)
        rows = [
            row
            for row in rows
            if (d := parse_date(row.get("expiry_date"))) and today <= d <= limit
        ]
        logger.info("تصفية التجديدات: %d ← %d خلال %d يوماً", before, len(rows), args.window)

    rows.sort(key=lambda row: parse_date(row.get("expiry_date")) or dt.date.max)
    try:
        export(rows, args.out)
    except RuntimeError as exc:
        logger.error(str(exc))
        raise SystemExit(1) from exc

    if CHECKPOINT.exists():
        CHECKPOINT.unlink()
        logger.info("اكتمل السحب — حُذف ملف الاستئناف.")


if __name__ == "__main__":
    main()
